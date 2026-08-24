"""缺陷统计桌面客户端。"""

from __future__ import annotations

import json
import io
import os
import queue
import sys
import threading
import webbrowser
from datetime import datetime
from pathlib import Path
from typing import Any, Callable

import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from PIL import Image, ImageTk

from analytics import filter_bugs, introducer_summary, summary_metrics, text
from api_client import (
    DEFAULT_BASE_URL,
    AuthenticationError,
    BugApiClient,
    BugClientError,
    PlanVersion,
)
from crm_client import (
    CRM_SEARCH_COLUMNS,
    BugDetail,
    BugDetailField,
    CrmApiClient,
    CrmAuthenticationError,
    CrmBugPage,
    CrmClientError,
    build_crm_bug_url,
)
from credential_store import load_credentials, save_credentials
from detail_utils import (
    extract_image_sources,
    rich_content_segments,
    rich_text_to_plain,
)
from retrospective import build_retrospective_markdown, retrospective_filename
from workload_client import (
    ALL_GROUP_MEMBERS,
    DEFAULT_WORKLOAD_GROUP,
    WorkloadApiClient,
    WorkloadContext,
    WorkloadPreview,
    WorkloadSubmitResult,
    build_workload_preview,
    read_workload_excel,
)


APP_TITLE = "缺陷统计客户端"
APP_VERSION = "v18.6"
WINDOW_TITLE = f"{APP_TITLE}  {APP_VERSION}"
APP_DIR = Path(__file__).resolve().parent
SETTINGS_DIR = Path(os.environ.get("APPDATA", APP_DIR)) / "BugStatisticsClient"
SETTINGS_PATH = SETTINGS_DIR / "settings.json"
CREDENTIALS_PATH = SETTINGS_DIR / "credentials.dat"
COLOR_BG = "#F3F6FB"
COLOR_SURFACE = "#FFFFFF"
COLOR_SURFACE_ALT = "#F7F9FC"
COLOR_SURFACE_RAISED = "#EDF2F8"
COLOR_BORDER = "#D7E0EB"
COLOR_TEXT = "#172033"
COLOR_MUTED = "#65748B"
COLOR_ACCENT = "#1677FF"
COLOR_ACCENT_HOVER = "#3B8CFF"
COLOR_CYAN = "#0EA5E9"
COLOR_SUCCESS = "#10B981"
COLOR_WARNING = "#F59E0B"
COLOR_DANGER = "#F43F5E"
DEFAULT_SETTINGS = {
    "base_url": DEFAULT_BASE_URL,
    "username": "T0423",
    "plan_version": "2026-0730",
    "introducer": "",
    "workload_group": DEFAULT_WORKLOAD_GROUP,
    "workload_developer": ALL_GROUP_MEMBERS,
    "workload_file": "",
}
CRM_TABLE_COLUMNS = CRM_SEARCH_COLUMNS

DETAIL_COLUMNS = [
    ("id", "缺陷编号", 165),
    ("_introducer", "缺陷引入人", 100),
    ("severity", "严重程度", 80),
    ("status", "状态", 80),
    ("module", "模块", 180),
    ("openedDate", "创建时间", 145),
    ("resolvedByName", "解决人", 100),
    ("title", "标题", 520),
]

SUMMARY_COLUMNS = [
    ("introducer", "缺陷引入人", 160),
    ("total", "总数", 80),
    ("severity_1", "1级", 70),
    ("severity_2", "2级", 70),
    ("severity_3", "3级", 70),
    ("severity_4", "4级", 70),
    ("closed", "已关闭", 80),
    ("open", "未关闭", 80),
]

WORKLOAD_COLUMNS = [
    ("excel_row", "Excel 行", 72),
    ("status", "状态", 82),
    ("developer", "责任人", 90),
    ("require_no", "需求编号", 135),
    ("task_name", "工作描述", 360),
    ("plan_start", "计划开始", 105),
    ("plan_finish", "计划完成", 105),
    ("requested_hours", "Excel 工时", 90),
    ("computed_hours", "计算工时", 90),
    ("message", "校验信息", 320),
]


class ImagePreviewDialog(tk.Toplevel):
    """可缩放、可滚动的原始图片查看器。"""

    def __init__(self, parent: tk.Misc, image_bytes: bytes) -> None:
        super().__init__(parent)
        self.title("图片查看")
        width = min(1200, self.winfo_screenwidth() - 80)
        height = min(850, self.winfo_screenheight() - 100)
        self.geometry(f"{width}x{height}")
        self.minsize(640, 480)
        self.transient(parent)
        self.configure(background=COLOR_BG)

        image = Image.open(io.BytesIO(image_bytes))
        image.load()
        self.original_image = image.copy()
        self.scale = 1.0
        self._photo: ImageTk.PhotoImage | None = None

        toolbar = ttk.Frame(self, padding=(10, 8))
        toolbar.pack(fill=tk.X)
        ttk.Button(
            toolbar,
            text="缩小",
            command=lambda: self._zoom(0.8),
            style="Secondary.TButton",
        ).pack(
            side=tk.LEFT
        )
        ttk.Button(
            toolbar,
            text="放大",
            command=lambda: self._zoom(1.25),
            style="Secondary.TButton",
        ).pack(
            side=tk.LEFT, padx=(6, 0)
        )
        ttk.Button(
            toolbar,
            text="100%",
            command=self._actual_size,
            style="Secondary.TButton",
        ).pack(
            side=tk.LEFT, padx=(6, 0)
        )
        ttk.Button(
            toolbar,
            text="适应窗口",
            command=self._fit_window,
            style="Accent.TButton",
        ).pack(
            side=tk.LEFT, padx=(6, 0)
        )
        self.scale_var = tk.StringVar(value="100%")
        ttk.Label(toolbar, textvariable=self.scale_var).pack(
            side=tk.LEFT, padx=(12, 0)
        )
        ttk.Label(
            toolbar,
            text=f"原始尺寸：{self.original_image.width} × {self.original_image.height}",
            style="Muted.TLabel",
        ).pack(side=tk.RIGHT)

        canvas_frame = ttk.Frame(self)
        canvas_frame.pack(fill=tk.BOTH, expand=True)
        self.canvas = tk.Canvas(
            canvas_frame,
            background="#030914",
            highlightthickness=0,
        )
        vertical = ttk.Scrollbar(
            canvas_frame,
            orient=tk.VERTICAL,
            command=self.canvas.yview,
        )
        horizontal = ttk.Scrollbar(
            canvas_frame,
            orient=tk.HORIZONTAL,
            command=self.canvas.xview,
        )
        self.canvas.configure(
            yscrollcommand=vertical.set,
            xscrollcommand=horizontal.set,
        )
        self.canvas.grid(row=0, column=0, sticky=tk.NSEW)
        vertical.grid(row=0, column=1, sticky=tk.NS)
        horizontal.grid(row=1, column=0, sticky=tk.EW)
        canvas_frame.rowconfigure(0, weight=1)
        canvas_frame.columnconfigure(0, weight=1)
        self.canvas.bind(
            "<Control-MouseWheel>",
            lambda event: self._zoom(1.25 if event.delta > 0 else 0.8),
        )
        self._render()

    def _render(self) -> None:
        width = max(1, round(self.original_image.width * self.scale))
        height = max(1, round(self.original_image.height * self.scale))
        if self.scale == 1.0:
            rendered = self.original_image
        else:
            rendered = self.original_image.resize(
                (width, height),
                Image.Resampling.LANCZOS,
            )
        self._photo = ImageTk.PhotoImage(rendered)
        self.canvas.delete("all")
        self.canvas.create_image(0, 0, image=self._photo, anchor=tk.NW)
        self.canvas.configure(scrollregion=(0, 0, width, height))
        self.scale_var.set(f"{self.scale * 100:.0f}%")

    def _zoom(self, factor: float) -> None:
        self.scale = min(3.0, max(0.1, self.scale * factor))
        self._render()

    def _actual_size(self) -> None:
        self.scale = 1.0
        self._render()

    def _fit_window(self) -> None:
        self.update_idletasks()
        available_width = max(1, self.canvas.winfo_width() - 8)
        available_height = max(1, self.canvas.winfo_height() - 8)
        self.scale = min(
            1.0,
            available_width / self.original_image.width,
            available_height / self.original_image.height,
        )
        self.scale = max(0.1, self.scale)
        self._render()


class BugDetailDialog(tk.Toplevel):
    """在统一滚动区域中展示全部 CRM 表单字段。"""

    def __init__(
        self,
        parent: tk.Misc,
        bug: dict[str, Any],
        detail: BugDetail,
        image_data: dict[str, bytes],
    ) -> None:
        super().__init__(parent)
        bug_number = detail.bug_number or text(bug.get("id"))
        title = detail.title or text(bug.get("title"))
        self.title(f"{bug_number} - 缺陷详情")
        self.geometry("1080x820")
        self.minsize(800, 600)
        self.transient(parent)
        self.configure(background=COLOR_BG)
        self.after_idle(self._maximize)

        self.bug_url = text(bug.get("bugUrl"))
        self.fields = list(detail.fields)
        if not self.fields:
            self.fields = [
                BugDetailField("description", "描述内容", detail.description_html, True),
                BugDetailField(
                    "DTS_RootCauseAnalysis",
                    "根因分析",
                    detail.root_cause_html,
                    True,
                ),
                BugDetailField(
                    "DTS_FixContent",
                    "修复内容",
                    detail.fix_content_html,
                    True,
                ),
            ]
        self.regular_fields = [
            field for field in self.fields if not field.is_rich_text
        ]
        self.rich_fields = [
            field for field in self.fields if field.is_rich_text
        ]
        self._photo_images: list[ImageTk.PhotoImage] = []
        self._image_labels: list[tk.Label] = []

        outer = ttk.Frame(self, padding=14)
        outer.pack(fill=tk.BOTH, expand=True)
        ttk.Label(
            outer,
            text=bug_number,
            font=("Microsoft YaHei UI", 13, "bold"),
        ).pack(anchor=tk.W)
        ttk.Label(
            outer,
            text=title,
            font=("Microsoft YaHei UI", 11),
            wraplength=1020,
        ).pack(anchor=tk.W, pady=(4, 10))

        content_frame = ttk.LabelFrame(outer, text="缺陷内容", padding=6)
        content_frame.pack(fill=tk.BOTH, expand=True)
        text_widget = tk.Text(
            content_frame,
            wrap=tk.WORD,
            font=("Microsoft YaHei UI", 10),
            relief=tk.FLAT,
            padx=10,
            pady=8,
            background=COLOR_SURFACE,
            foreground=COLOR_TEXT,
            insertbackground=COLOR_TEXT,
            selectbackground=COLOR_ACCENT,
            selectforeground="#FFFFFF",
            highlightbackground=COLOR_BORDER,
            highlightcolor=COLOR_CYAN,
            highlightthickness=1,
        )
        scrollbar = ttk.Scrollbar(
            content_frame,
            orient=tk.VERTICAL,
            command=text_widget.yview,
        )
        text_widget.configure(yscrollcommand=scrollbar.set)
        text_widget.tag_configure(
            "section_heading",
            font=("Microsoft YaHei UI", 11, "bold"),
            foreground=COLOR_CYAN,
            spacing1=10,
            spacing3=6,
        )
        text_widget.tag_configure(
            "field_label",
            font=("Microsoft YaHei UI", 10, "bold"),
            foreground="#40536B",
        )
        text_widget.insert(tk.END, "全部表单字段\n", "section_heading")
        for field in self.regular_fields:
            text_widget.insert(tk.END, field.label + "：", "field_label")
            text_widget.insert(tk.END, field.value + "\n")
        for field in self.rich_fields:
            text_widget.insert(tk.END, "\n")
            text_widget.insert(tk.END, field.label + "\n", "section_heading")
            self._insert_rich_content(text_widget, field.value, image_data)
            text_widget.insert(tk.END, "\n")
        text_widget.configure(state=tk.DISABLED)
        text_widget.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        buttons = ttk.Frame(outer)
        buttons.pack(fill=tk.X, pady=(10, 0))
        ttk.Label(
            buttons,
            text="点击图片可查看原始尺寸；Ctrl + 鼠标滚轮可缩放原图",
            style="Muted.TLabel",
        ).pack(side=tk.LEFT)
        ttk.Button(
            buttons,
            text="复制全部",
            command=self._copy_all,
            style="Secondary.TButton",
        ).pack(
            side=tk.RIGHT
        )
        ttk.Button(
            buttons,
            text="关闭",
            command=self.destroy,
            style="Accent.TButton",
        ).pack(
            side=tk.RIGHT, padx=(0, 8)
        )
        if self.bug_url:
            ttk.Button(
                buttons,
                text="在 CRM 中打开",
                command=lambda: webbrowser.open(self.bug_url),
                style="Secondary.TButton",
            ).pack(side=tk.RIGHT, padx=(0, 8))

    def _maximize(self) -> None:
        """窗口首次显示时自动最大化。"""
        try:
            self.state("zoomed")
        except tk.TclError:
            # 不支持 zoomed 状态的平台仍使用上面的默认窗口尺寸。
            pass

    def _insert_rich_content(
        self,
        widget: tk.Text,
        rich_content: str,
        image_data: dict[str, bytes],
    ) -> None:
        for kind, content in rich_content_segments(rich_content):
            if kind == "text":
                widget.insert(tk.END, content)
                continue
            data = image_data.get(content)
            if not data:
                widget.insert(tk.END, "[图片加载失败]")
                continue
            try:
                original = Image.open(io.BytesIO(data))
                original.load()
                thumbnail = original.copy()
                thumbnail.thumbnail((900, 900), Image.Resampling.LANCZOS)
                photo = ImageTk.PhotoImage(thumbnail)
            except Exception:
                widget.insert(tk.END, "[图片格式无法识别]")
                continue
            if widget.index(tk.END) != "1.0":
                widget.insert(tk.END, "\n")
            image_label = tk.Label(
                widget,
                image=photo,
                cursor="hand2",
                background=COLOR_SURFACE,
                borderwidth=1,
                relief=tk.SOLID,
            )
            image_label.bind(
                "<Button-1>",
                lambda _event, raw=data: self._open_image(raw),
            )
            widget.window_create(tk.END, window=image_label)
            widget.insert(tk.END, "\n")
            self._photo_images.append(photo)
            self._image_labels.append(image_label)

    def _open_image(self, image_bytes: bytes) -> ImagePreviewDialog:
        return ImagePreviewDialog(self, image_bytes)

    def _copy_all(self) -> None:
        content = "\n".join(
            (
                f"【{field.label}】\n{rich_text_to_plain(field.value)}"
                if field.is_rich_text
                else f"{field.label}：{field.value}"
            )
            for field in self.fields
        )
        self.clipboard_clear()
        self.clipboard_append(content)


class BugStatisticsApp(tk.Tk):
    def __init__(self) -> None:
        super().__init__()
        self.title(WINDOW_TITLE)
        self.geometry("1380x860")
        self.minsize(1100, 680)
        self.configure(background=COLOR_BG)

        self.settings = self._load_settings()
        self.credentials = load_credentials(CREDENTIALS_PATH)
        self.client = BugApiClient(self.settings["base_url"])
        self.workload_api = WorkloadApiClient(self.client)
        self.crm_client = CrmApiClient()
        self.all_bugs: list[dict[str, Any]] = []
        self.filtered_bugs: list[dict[str, Any]] = []
        self.crm_bug_page: CrmBugPage | None = None
        self.worker_messages: queue.Queue[tuple[str, Any]] = queue.Queue()
        self.busy = False
        self.logged_in = False
        self.plan_versions: list[PlanVersion] = []
        self.workload_context: WorkloadContext | None = None
        self.workload_preview: WorkloadPreview | None = None
        self.performance_display_name = ""

        self._configure_style()
        self._build_ui()
        self.protocol("WM_DELETE_WINDOW", self._on_close)
        self.after(100, self._poll_worker_messages)

    def _load_settings(self) -> dict[str, str]:
        settings = dict(DEFAULT_SETTINGS)
        try:
            stored = json.loads(SETTINGS_PATH.read_text(encoding="utf-8"))
            if isinstance(stored, dict):
                for key in settings:
                    if isinstance(stored.get(key), str):
                        settings[key] = stored[key]
        except (OSError, ValueError):
            pass
        return settings

    def _save_settings(self) -> None:
        settings = {
            "base_url": self.settings["base_url"],
            "username": self.username_var.get().strip(),
            "plan_version": self.plan_var.get().strip(),
            "introducer": self.introducer_var.get().strip(),
            "workload_group": self.workload_group_var.get().strip(),
            "workload_developer": self.workload_developer_var.get().strip(),
            "workload_file": self.workload_file_var.get().strip(),
        }
        try:
            SETTINGS_DIR.mkdir(parents=True, exist_ok=True)
            SETTINGS_PATH.write_text(
                json.dumps(settings, ensure_ascii=False, indent=2),
                encoding="utf-8",
            )
        except OSError:
            pass

    def _configure_style(self) -> None:
        style = ttk.Style(self)
        if "clam" in style.theme_names():
            style.theme_use("clam")

        default_font = ("Microsoft YaHei UI", 10)
        style.configure(
            ".",
            background=COLOR_BG,
            foreground=COLOR_TEXT,
            font=default_font,
        )
        style.configure("TFrame", background=COLOR_BG)
        style.configure("Header.TFrame", background=COLOR_BG)
        style.configure("Card.TFrame", background=COLOR_SURFACE)
        style.configure(
            "StatusBar.TFrame",
            background=COLOR_SURFACE,
            borderwidth=1,
            relief=tk.SOLID,
        )
        style.configure(
            "TLabel",
            background=COLOR_BG,
            foreground=COLOR_TEXT,
        )
        style.configure(
            "Title.TLabel",
            background=COLOR_BG,
            foreground=COLOR_TEXT,
            font=("Microsoft YaHei UI", 20, "bold"),
        )
        style.configure(
            "SubTitle.TLabel",
            background=COLOR_BG,
            foreground=COLOR_MUTED,
            font=("Microsoft YaHei UI", 10),
        )
        style.configure(
            "Eyebrow.TLabel",
            background=COLOR_BG,
            foreground=COLOR_CYAN,
            font=("Microsoft YaHei UI", 8, "bold"),
        )
        style.configure(
            "Muted.TLabel",
            background=COLOR_BG,
            foreground=COLOR_MUTED,
        )
        style.configure(
            "Card.TLabel",
            background=COLOR_SURFACE,
            foreground=COLOR_TEXT,
        )
        style.configure(
            "CardMuted.TLabel",
            background=COLOR_SURFACE,
            foreground=COLOR_MUTED,
        )
        style.configure(
            "LoginStatus.TLabel",
            background="#FFF7E6",
            foreground="#B86A00",
            padding=(12, 7),
            font=("Microsoft YaHei UI", 9, "bold"),
        )
        style.configure(
            "Connected.TLabel",
            background="#E8FBF3",
            foreground="#078B68",
            padding=(12, 7),
            font=("Microsoft YaHei UI", 9, "bold"),
        )
        style.configure(
            "Status.TLabel",
            background=COLOR_SURFACE,
            foreground=COLOR_MUTED,
        )

        style.configure(
            "TLabelframe",
            background=COLOR_SURFACE,
            bordercolor=COLOR_BORDER,
            lightcolor=COLOR_BORDER,
            darkcolor=COLOR_BORDER,
            borderwidth=1,
            relief=tk.SOLID,
        )
        style.configure(
            "TLabelframe.Label",
            background=COLOR_SURFACE,
            foreground=COLOR_ACCENT,
            font=("Microsoft YaHei UI", 9, "bold"),
        )

        style.configure(
            "TEntry",
            fieldbackground=COLOR_SURFACE_ALT,
            foreground=COLOR_TEXT,
            insertcolor=COLOR_TEXT,
            bordercolor=COLOR_BORDER,
            lightcolor=COLOR_BORDER,
            darkcolor=COLOR_BORDER,
            padding=(8, 6),
            relief=tk.FLAT,
        )
        style.map(
            "TEntry",
            bordercolor=[("focus", COLOR_CYAN)],
            lightcolor=[("focus", COLOR_CYAN)],
            darkcolor=[("focus", COLOR_CYAN)],
            fieldbackground=[("disabled", COLOR_SURFACE)],
        )
        style.configure(
            "TCombobox",
            fieldbackground=COLOR_SURFACE_ALT,
            background=COLOR_SURFACE_ALT,
            foreground=COLOR_TEXT,
            arrowcolor=COLOR_CYAN,
            bordercolor=COLOR_BORDER,
            lightcolor=COLOR_BORDER,
            darkcolor=COLOR_BORDER,
            padding=(8, 5),
        )
        style.map(
            "TCombobox",
            fieldbackground=[
                ("readonly", COLOR_SURFACE_ALT),
                ("focus", COLOR_SURFACE_ALT),
            ],
            foreground=[("readonly", COLOR_TEXT)],
            selectbackground=[("readonly", COLOR_SURFACE_ALT)],
            selectforeground=[("readonly", COLOR_TEXT)],
            bordercolor=[("focus", COLOR_CYAN)],
        )

        style.configure(
            "TButton",
            background=COLOR_SURFACE_RAISED,
            foreground=COLOR_TEXT,
            borderwidth=0,
            focusthickness=0,
            padding=(12, 7),
            relief=tk.FLAT,
            font=("Microsoft YaHei UI", 9, "bold"),
        )
        style.map(
            "TButton",
            background=[
                ("pressed", "#D4E5FA"),
                ("active", "#E1ECFA"),
                ("disabled", "#E9EEF5"),
            ],
            foreground=[("disabled", "#98A5B5")],
        )
        style.configure(
            "Accent.TButton",
            background=COLOR_ACCENT,
            foreground="#FFFFFF",
        )
        style.map(
            "Accent.TButton",
            background=[
                ("pressed", "#1766D7"),
                ("active", COLOR_ACCENT_HOVER),
                ("disabled", "#A9C9F5"),
            ],
            foreground=[("disabled", "#F3F7FD")],
        )
        style.configure(
            "Success.TButton",
            background="#147D68",
            foreground="#FFFFFF",
        )
        style.map(
            "Success.TButton",
            background=[
                ("pressed", "#0E6655"),
                ("active", "#19967C"),
                ("disabled", "#A7D7CA"),
            ],
        )
        style.configure(
            "Secondary.TButton",
            background=COLOR_SURFACE_RAISED,
            foreground="#40536B",
        )
        style.configure(
            "Icon.TButton",
            background=COLOR_SURFACE_RAISED,
            foreground=COLOR_CYAN,
            padding=(5, 5),
        )
        style.configure(
            "Markdown.TButton",
            background="#7357E8",
            foreground="#FFFFFF",
        )
        style.map(
            "Markdown.TButton",
            background=[
                ("pressed", "#5A3FC6"),
                ("active", "#876EF0"),
                ("disabled", "#C5BDEB"),
            ],
            foreground=[("disabled", "#F7F5FF")],
        )

        style.configure(
            "MetricValue.TLabel",
            background=COLOR_SURFACE,
            foreground=COLOR_CYAN,
            font=("Microsoft YaHei UI", 21, "bold"),
        )
        style.configure(
            "MetricName.TLabel",
            background=COLOR_SURFACE,
            foreground=COLOR_MUTED,
            font=("Microsoft YaHei UI", 9),
        )

        style.configure(
            "Content.TNotebook",
            background=COLOR_BG,
            borderwidth=0,
            tabmargins=0,
        )
        style.layout("Content.TNotebook.Tab", [])
        style.configure(
            "NavActive.TButton",
            background="#E8F1FF",
            foreground=COLOR_ACCENT,
            borderwidth=0,
            padding=(22, 10),
            font=("Microsoft YaHei UI", 10, "bold"),
        )
        style.map(
            "NavActive.TButton",
            background=[("pressed", "#DCEAFF"), ("active", "#E2EDFF")],
            foreground=[("active", COLOR_ACCENT)],
        )
        style.configure(
            "NavInactive.TButton",
            background=COLOR_SURFACE,
            foreground="#52637A",
            borderwidth=0,
            padding=(22, 10),
            font=("Microsoft YaHei UI", 10, "bold"),
        )
        style.map(
            "NavInactive.TButton",
            background=[("pressed", "#EDF3FA"), ("active", "#F2F6FB")],
            foreground=[("active", COLOR_TEXT)],
        )

        style.configure(
            "Treeview",
            background=COLOR_SURFACE,
            fieldbackground=COLOR_SURFACE,
            foreground=COLOR_TEXT,
            borderwidth=0,
            relief=tk.FLAT,
            rowheight=32,
            font=("Microsoft YaHei UI", 9),
        )
        style.map(
            "Treeview",
            background=[("selected", COLOR_ACCENT)],
            foreground=[("selected", "#FFFFFF")],
        )
        style.configure(
            "Treeview.Heading",
            background=COLOR_SURFACE_RAISED,
            foreground="#40536B",
            borderwidth=0,
            relief=tk.FLAT,
            padding=(8, 8),
            font=("Microsoft YaHei UI", 9, "bold"),
        )
        style.map(
            "Treeview.Heading",
            background=[("active", "#DEEAF7")],
            foreground=[("active", COLOR_TEXT)],
        )
        for scrollbar_style in (
            "TScrollbar",
            "Vertical.TScrollbar",
            "Horizontal.TScrollbar",
        ):
            style.configure(
                scrollbar_style,
                background=COLOR_SURFACE_RAISED,
                troughcolor=COLOR_SURFACE,
                bordercolor=COLOR_BORDER,
                lightcolor=COLOR_SURFACE_RAISED,
                darkcolor=COLOR_SURFACE_RAISED,
                arrowcolor=COLOR_CYAN,
                relief=tk.FLAT,
                borderwidth=0,
            )
            style.map(
                scrollbar_style,
                background=[
                    ("pressed", COLOR_ACCENT),
                    ("active", "#DCE8F6"),
                ],
                arrowcolor=[("active", COLOR_ACCENT)],
            )
        style.configure(
            "Horizontal.TProgressbar",
            background=COLOR_CYAN,
            troughcolor=COLOR_SURFACE_RAISED,
            bordercolor=COLOR_SURFACE,
            lightcolor=COLOR_CYAN,
            darkcolor=COLOR_CYAN,
        )
        self.option_add("*TCombobox*Listbox.background", COLOR_SURFACE_ALT)
        self.option_add("*TCombobox*Listbox.foreground", COLOR_TEXT)
        self.option_add("*TCombobox*Listbox.selectBackground", COLOR_ACCENT)
        self.option_add("*TCombobox*Listbox.selectForeground", "#FFFFFF")

    def _build_ui(self) -> None:
        outer = ttk.Frame(self, padding=18)
        outer.pack(fill=tk.BOTH, expand=True)

        header = ttk.Frame(outer, style="Header.TFrame")
        header.pack(fill=tk.X, pady=(0, 14))
        tk.Frame(header, width=5, background=COLOR_CYAN).pack(
            side=tk.LEFT,
            fill=tk.Y,
            padx=(0, 12),
        )
        title_block = ttk.Frame(header, style="Header.TFrame")
        title_block.pack(side=tk.LEFT)
        ttk.Label(
            title_block,
            text="BUG INTELLIGENCE CONSOLE",
            style="Eyebrow.TLabel",
        ).pack(anchor=tk.W)
        title_line = ttk.Frame(title_block, style="Header.TFrame")
        title_line.pack(anchor=tk.W)
        ttk.Label(
            title_line,
            text=APP_TITLE,
            style="Title.TLabel",
        ).pack(side=tk.LEFT)
        ttk.Label(
            title_line,
            text="绩效分析  /  CRM 检索  /  批量录入",
            style="SubTitle.TLabel",
        ).pack(side=tk.LEFT, padx=(16, 0), pady=(8, 0))
        self.login_status_var = tk.StringVar(value="未登录")
        self.login_status_label = ttk.Label(
            header,
            textvariable=self.login_status_var,
            style="LoginStatus.TLabel",
        )
        self.login_status_label.pack(
            side=tk.RIGHT,
            pady=(8, 0),
        )

        self._build_login_controls(outer)
        self._build_tables(outer)

        status_frame = ttk.Frame(outer, style="StatusBar.TFrame", padding=(10, 7))
        status_frame.pack(fill=tk.X, pady=(8, 0))
        self.status_var = tk.StringVar(value="请输入绩效密码和 CRM 密码后登录")
        ttk.Label(
            status_frame,
            textvariable=self.status_var,
            style="Status.TLabel",
        ).pack(side=tk.LEFT)
        self.progress = ttk.Progressbar(status_frame, mode="indeterminate", length=160)
        self.progress.pack(side=tk.RIGHT)
        self._build_loading_windows()

    def _build_loading_windows(self) -> None:
        self.loading_dimmer = tk.Toplevel(self)
        self.loading_dimmer.withdraw()
        self.loading_dimmer.overrideredirect(True)
        self.loading_dimmer.transient(self)
        self.loading_dimmer.configure(background=COLOR_TEXT, cursor="wait")
        self.loading_dimmer.attributes("-alpha", 0.22)
        self.loading_dimmer.bind("<Button>", lambda _event: "break")
        self.loading_dimmer.bind("<MouseWheel>", lambda _event: "break")

        self.loading_dialog = tk.Toplevel(self)
        self.loading_dialog.withdraw()
        self.loading_dialog.overrideredirect(True)
        self.loading_dialog.transient(self)
        self.loading_dialog.configure(background=COLOR_BORDER, cursor="wait")
        self.loading_dialog.bind("<Button>", lambda _event: "break")
        self.loading_dialog.bind("<Key>", lambda _event: "break")

        loading_card = tk.Frame(
            self.loading_dialog,
            background=COLOR_SURFACE,
            padx=34,
            pady=26,
        )
        loading_card.pack(padx=1, pady=1)
        tk.Label(
            loading_card,
            text="正在处理，请稍候…",
            background=COLOR_SURFACE,
            foreground=COLOR_TEXT,
            font=("Microsoft YaHei UI", 13, "bold"),
        ).pack()
        tk.Label(
            loading_card,
            textvariable=self.status_var,
            background=COLOR_SURFACE,
            foreground=COLOR_MUTED,
            font=("Microsoft YaHei UI", 9),
        ).pack(pady=(8, 14))
        self.loading_progress = ttk.Progressbar(
            loading_card,
            mode="indeterminate",
            length=280,
        )
        self.loading_progress.pack()
        self.bind("<Configure>", self._position_loading_windows, add="+")

    def _position_loading_windows(self, _event: tk.Event | None = None) -> None:
        if not self.busy:
            return
        root_x = self.winfo_rootx()
        root_y = self.winfo_rooty()
        root_width = max(1, self.winfo_width())
        root_height = max(1, self.winfo_height())
        self.loading_dimmer.geometry(
            f"{root_width}x{root_height}+{root_x}+{root_y}"
        )
        dialog_width = max(350, self.loading_dialog.winfo_reqwidth())
        dialog_height = max(140, self.loading_dialog.winfo_reqheight())
        dialog_x = root_x + (root_width - dialog_width) // 2
        dialog_y = root_y + (root_height - dialog_height) // 2
        self.loading_dialog.geometry(
            f"{dialog_width}x{dialog_height}+{dialog_x}+{dialog_y}"
        )

    def _show_loading(self) -> None:
        self.update_idletasks()
        self._position_loading_windows()
        self.loading_dimmer.deiconify()
        self.loading_dimmer.lift()
        self.loading_dialog.deiconify()
        self.loading_dialog.lift()
        self.loading_dialog.grab_set()
        self.loading_dialog.focus_force()
        self.loading_progress.start(10)

    def _hide_loading(self) -> None:
        self.loading_progress.stop()
        try:
            if self.grab_current() == self.loading_dialog:
                self.loading_dialog.grab_release()
        except tk.TclError:
            pass
        self.loading_dialog.withdraw()
        self.loading_dimmer.withdraw()

    def _build_login_controls(self, parent: ttk.Frame) -> None:
        login_box = ttk.LabelFrame(parent, text="登录", padding=10)
        login_box.pack(fill=tk.X, pady=(0, 10))

        saved_username = self.credentials.get("username", "")
        username = saved_username or self.settings["username"]
        saved_for_username = saved_username == username
        self.username_var = tk.StringVar(value=username)
        self.password_var = tk.StringVar(
            value=self.credentials.get("password", "") if saved_for_username else ""
        )
        self.crm_password_var = tk.StringVar(
            value=(
                self.credentials.get("crm_password", "")
                if saved_for_username
                else ""
            )
        )
        ttk.Label(login_box, text="用户名", style="Card.TLabel").grid(
            row=0, column=0, padx=(0, 6)
        )
        username_entry = ttk.Entry(
            login_box, textvariable=self.username_var, width=18
        )
        username_entry.grid(row=0, column=1, padx=(0, 14))
        ttk.Label(login_box, text="绩效密码", style="Card.TLabel").grid(
            row=0, column=2, padx=(0, 6)
        )
        password_entry = ttk.Entry(
            login_box, textvariable=self.password_var, show="●", width=18
        )
        password_entry.grid(row=0, column=3, padx=(0, 4))
        password_eye_button = ttk.Button(
            login_box,
            text="👁",
            width=2,
            takefocus=False,
            style="Icon.TButton",
        )
        password_eye_button.configure(
            command=lambda: self._toggle_password_visibility(
                password_entry,
                password_eye_button,
            )
        )
        password_eye_button.grid(row=0, column=4, padx=(0, 14))
        ttk.Label(login_box, text="CRM 密码", style="Card.TLabel").grid(
            row=0, column=5, padx=(0, 6)
        )
        crm_password_entry = ttk.Entry(
            login_box,
            textvariable=self.crm_password_var,
            show="●",
            width=18,
        )
        crm_password_entry.grid(row=0, column=6, padx=(0, 4))
        crm_password_eye_button = ttk.Button(
            login_box,
            text="👁",
            width=2,
            takefocus=False,
            style="Icon.TButton",
        )
        crm_password_eye_button.configure(
            command=lambda: self._toggle_password_visibility(
                crm_password_entry,
                crm_password_eye_button,
            )
        )
        crm_password_eye_button.grid(row=0, column=7, padx=(0, 14))
        self.login_button = ttk.Button(
            login_box,
            text="登录",
            command=self._login,
            style="Accent.TButton",
        )
        self.login_button.grid(row=0, column=8, padx=(0, 8))
        ttk.Label(
            login_box,
            text="登录成功后由 Windows 当前账户加密记住",
            style="CardMuted.TLabel",
        ).grid(row=0, column=9, sticky=tk.W)
        login_box.columnconfigure(9, weight=1)
        password_entry.bind("<Return>", lambda _event: self._login())
        crm_password_entry.bind("<Return>", lambda _event: self._login())

    def _build_performance_controls(self, parent: ttk.Frame) -> None:
        query_box = ttk.LabelFrame(parent, text="绩效查询与筛选", padding=10)
        self.performance_query_box = query_box
        query_box.pack(fill=tk.X, pady=(0, 10))
        self.plan_var = tk.StringVar(value=self.settings["plan_version"])
        self.introducer_var = tk.StringVar(value=self.settings["introducer"])
        self.keyword_var = tk.StringVar()

        ttk.Label(query_box, text="计划版本", style="Card.TLabel").grid(
            row=0, column=0, padx=(0, 6)
        )
        self.plan_combo = ttk.Combobox(
            query_box,
            textvariable=self.plan_var,
            width=16,
            state="readonly",
        )
        self.plan_combo.grid(
            row=0, column=1, padx=(0, 14)
        )
        ttk.Label(query_box, text="缺陷引入人", style="Card.TLabel").grid(
            row=0, column=2, padx=(0, 6)
        )
        introducer_entry = ttk.Entry(
            query_box, textvariable=self.introducer_var, width=15
        )
        introducer_entry.grid(row=0, column=3, padx=(0, 14))
        ttk.Label(query_box, text="关键字", style="Card.TLabel").grid(
            row=0, column=4, padx=(0, 6)
        )
        keyword_entry = ttk.Entry(
            query_box, textvariable=self.keyword_var, width=30
        )
        keyword_entry.grid(row=0, column=5, padx=(0, 14), sticky=tk.EW)
        query_box.columnconfigure(5, weight=1)

        self.query_button = ttk.Button(
            query_box,
            text="查询接口",
            command=self._query,
            style="Accent.TButton",
        )
        self.query_button.grid(row=0, column=6, padx=(0, 8))
        self.filter_button = ttk.Button(
            query_box,
            text="应用筛选",
            command=self._apply_filter,
            style="Secondary.TButton",
        )
        self.filter_button.grid(row=0, column=7, padx=(0, 8))
        self.clear_button = ttk.Button(
            query_box,
            text="清空筛选",
            command=self._clear_filter,
            style="Secondary.TButton",
        )
        self.clear_button.grid(row=0, column=8, padx=(0, 8))
        self.export_button = ttk.Button(
            query_box,
            text="导出 Excel",
            command=self._export_excel,
            style="Success.TButton",
        )
        self.export_button.grid(row=0, column=9, padx=(0, 8))
        self.export_md_button = ttk.Button(
            query_box,
            text="导出回溯 MD",
            command=self._export_retrospective_md,
            style="Markdown.TButton",
        )
        self.export_md_button.grid(row=0, column=10)

        introducer_entry.bind("<Return>", lambda _event: self._apply_filter())
        keyword_entry.bind("<Return>", lambda _event: self._apply_filter())

    def _build_workload_controls(self, parent: ttk.Frame) -> None:
        controls = ttk.LabelFrame(parent, text="绩效工作量批量录入", padding=10)
        controls.grid(row=0, column=0, sticky=tk.EW, pady=(0, 8))
        controls.columnconfigure(9, weight=1)

        self.workload_plan_var = tk.StringVar(value=self.settings["plan_version"])
        self.workload_group_var = tk.StringVar(
            value=self.settings.get("workload_group", DEFAULT_WORKLOAD_GROUP)
        )
        self.workload_developer_var = tk.StringVar(
            value=self.settings.get("workload_developer", ALL_GROUP_MEMBERS)
            or ALL_GROUP_MEMBERS
        )
        self.workload_file_var = tk.StringVar(
            value=self.settings.get("workload_file", "")
        )

        ttk.Label(controls, text="计划版本", style="Card.TLabel").grid(
            row=0, column=0, padx=(0, 6)
        )
        self.workload_plan_combo = ttk.Combobox(
            controls,
            textvariable=self.workload_plan_var,
            width=15,
            state="readonly",
        )
        self.workload_plan_combo.grid(row=0, column=1, padx=(0, 12))

        ttk.Label(controls, text="录入分组", style="Card.TLabel").grid(
            row=0, column=2, padx=(0, 6)
        )
        self.workload_group_combo = ttk.Combobox(
            controls,
            textvariable=self.workload_group_var,
            width=17,
            state="readonly",
        )
        self.workload_group_combo.grid(row=0, column=3, padx=(0, 12))
        self.workload_group_combo.bind(
            "<<ComboboxSelected>>", self._workload_group_changed
        )

        ttk.Label(controls, text="责任人筛选", style="Card.TLabel").grid(
            row=0, column=4, padx=(0, 6)
        )
        self.workload_developer_combo = ttk.Combobox(
            controls,
            textvariable=self.workload_developer_var,
            values=[ALL_GROUP_MEMBERS],
            width=13,
            state="readonly",
        )
        self.workload_developer_combo.grid(row=0, column=5, padx=(0, 12))
        self.workload_developer_combo.bind(
            "<<ComboboxSelected>>", lambda _event: self._clear_workload_preview()
        )

        ttk.Label(controls, text="Excel", style="Card.TLabel").grid(
            row=0, column=6, padx=(0, 6)
        )
        workload_file_entry = ttk.Entry(
            controls,
            textvariable=self.workload_file_var,
            state="readonly",
            width=38,
        )
        workload_file_entry.grid(row=0, column=7, sticky=tk.EW, padx=(0, 6))
        controls.columnconfigure(7, weight=1)
        self.workload_browse_button = ttk.Button(
            controls,
            text="选择文件",
            command=self._choose_workload_file,
            style="Secondary.TButton",
        )
        self.workload_browse_button.grid(row=0, column=8, padx=(0, 8))
        self.workload_preview_button = ttk.Button(
            controls,
            text="加载并预览",
            command=self._preview_workload,
            style="Accent.TButton",
        )
        self.workload_preview_button.grid(row=0, column=9, sticky=tk.E)

        ttk.Label(
            controls,
            text="只读取 .xlsx；选择“全部组员”时，仅导入当前分组成员的数据。提交前会再次确认。",
            style="CardMuted.TLabel",
        ).grid(row=1, column=0, columnspan=10, sticky=tk.W, pady=(8, 0))

    def _build_workload_tab(self, parent: ttk.Frame) -> None:
        self._build_workload_controls(parent)
        preview_frame = ttk.Frame(parent)
        preview_frame.grid(row=1, column=0, sticky=tk.NSEW)
        parent.rowconfigure(1, weight=1)
        parent.columnconfigure(0, weight=1)
        self.workload_tree = self._make_tree(preview_frame, WORKLOAD_COLUMNS)
        workload_text_columns = {"task_name", "message"}
        for key, _label, _width in WORKLOAD_COLUMNS:
            alignment = tk.W if key in workload_text_columns else tk.CENTER
            self.workload_tree.heading(key, anchor=alignment)
            self.workload_tree.column(
                key,
                anchor=alignment,
                stretch=key == "message",
            )
        self.workload_tree.tag_configure("valid", background="#ECFDF5")
        self.workload_tree.tag_configure("warning", background="#FFFBEB")
        self.workload_tree.tag_configure("error", background="#FFF1F2")
        self.workload_tree.tag_configure("duplicate", background="#F1F5F9")

        footer = ttk.Frame(parent, style="Card.TFrame", padding=(8, 8))
        footer.grid(row=2, column=0, sticky=tk.EW, pady=(8, 0))
        self.workload_summary_var = tk.StringVar(
            value="请选择 Excel，加载后先预览校验结果"
        )
        ttk.Label(
            footer,
            textvariable=self.workload_summary_var,
            style="CardMuted.TLabel",
        ).pack(side=tk.LEFT)
        self.workload_clear_button = ttk.Button(
            footer,
            text="清空预览",
            command=self._clear_workload_preview,
            style="Secondary.TButton",
        )
        self.workload_clear_button.pack(side=tk.RIGHT, padx=(8, 0))
        self.workload_submit_button = ttk.Button(
            footer,
            text="确认批量提交",
            command=self._submit_workload,
            state=tk.DISABLED,
            style="Success.TButton",
        )
        self.workload_submit_button.pack(side=tk.RIGHT)

    @staticmethod
    def _toggle_password_visibility(
        entry: ttk.Entry,
        button: ttk.Button,
    ) -> None:
        is_visible = entry.cget("show") == ""
        entry.configure(show="●" if is_visible else "")
        button.configure(text="👁" if is_visible else "🙈")
        entry.focus_set()
        entry.icursor(tk.END)

    def _build_metrics(self, parent: ttk.Frame) -> None:
        metrics_box = ttk.Frame(parent)
        self.performance_metrics_box = metrics_box
        metrics_box.pack(fill=tk.X, pady=(0, 10))
        metric_defs = [
            ("total", "筛选结果", COLOR_CYAN),
            ("severity_1", "1级", COLOR_DANGER),
            ("severity_2", "2级", COLOR_WARNING),
            ("severity_3", "3级", "#A78BFA"),
            ("severity_4", "4级", "#38BDF8"),
            ("closed", "已关闭", COLOR_SUCCESS),
            ("open", "未关闭", "#FB923C"),
        ]
        self.metric_vars: dict[str, tk.StringVar] = {}
        for index, (key, label, accent_color) in enumerate(metric_defs):
            card = tk.Frame(
                metrics_box,
                background=COLOR_SURFACE,
                highlightbackground=COLOR_BORDER,
                highlightcolor=accent_color,
                highlightthickness=1,
            )
            card.grid(
                row=0,
                column=index,
                padx=(0 if index == 0 else 7, 0),
                sticky=tk.EW,
            )
            metrics_box.columnconfigure(index, weight=1)
            tk.Frame(
                card,
                height=3,
                background=accent_color,
            ).pack(fill=tk.X)
            card_content = tk.Frame(card, background=COLOR_SURFACE)
            card_content.pack(fill=tk.BOTH, expand=True, padx=12, pady=(7, 8))
            value_var = tk.StringVar(value="0")
            self.metric_vars[key] = value_var
            tk.Label(
                card_content,
                textvariable=value_var,
                background=COLOR_SURFACE,
                foreground=accent_color,
                font=("Microsoft YaHei UI", 21, "bold"),
            ).pack()
            tk.Label(
                card_content,
                text=label,
                background=COLOR_SURFACE,
                foreground=COLOR_MUTED,
                font=("Microsoft YaHei UI", 9),
            ).pack()

    def _build_tables(self, parent: ttk.Frame) -> None:
        nav_shell = tk.Frame(
            parent,
            background=COLOR_SURFACE,
            highlightbackground=COLOR_BORDER,
            highlightcolor=COLOR_BORDER,
            highlightthickness=1,
        )
        nav_shell.pack(anchor=tk.W, pady=(0, 10))
        nav_content = ttk.Frame(nav_shell, style="Card.TFrame")
        nav_content.pack(padx=5, pady=5)
        self.performance_nav_button = ttk.Button(
            nav_content,
            text="▦  绩效缺陷明细",
            command=lambda: self._select_main_tab("performance"),
            style="NavActive.TButton",
        )
        self.performance_nav_button.pack(side=tk.LEFT)
        self.crm_nav_button = ttk.Button(
            nav_content,
            text="⌕  CRM 缺陷查询",
            command=lambda: self._select_main_tab("crm"),
            style="NavInactive.TButton",
        )
        self.crm_nav_button.pack(side=tk.LEFT, padx=(4, 0))
        self.workload_nav_button = ttk.Button(
            nav_content,
            text="▤  绩效录入",
            command=lambda: self._select_main_tab("workload"),
            style="NavInactive.TButton",
        )
        self.workload_nav_button.pack(side=tk.LEFT, padx=(4, 0))

        self.notebook = ttk.Notebook(parent, style="Content.TNotebook")
        self.notebook.pack(fill=tk.BOTH, expand=True)

        self.performance_detail_tab = ttk.Frame(self.notebook, padding=6)
        self.crm_tab = ttk.Frame(self.notebook, padding=6)
        self.workload_tab = ttk.Frame(self.notebook, padding=6)
        self.notebook.add(self.performance_detail_tab, text="绩效缺陷明细")
        self.notebook.add(self.crm_tab, text="CRM 缺陷查询")
        self.notebook.add(self.workload_tab, text="绩效录入")

        self._build_performance_controls(self.performance_detail_tab)
        self._build_metrics(self.performance_detail_tab)
        detail_tree_frame = ttk.Frame(self.performance_detail_tab)
        detail_tree_frame.pack(fill=tk.BOTH, expand=True)
        self.detail_tree = self._make_tree(detail_tree_frame, DETAIL_COLUMNS)
        self.detail_number_widgets: dict[
            str, tuple[tk.Frame, tk.Label, tk.Canvas]
        ] = {}
        self.detail_tree.bind(
            "<<TreeviewScroll>>",
            lambda _event: self.after_idle(self._position_detail_number_widgets),
        )
        self.detail_tree.bind(
            "<<TreeviewSelect>>",
            lambda _event: self.after_idle(self._position_detail_number_widgets),
        )
        self.detail_tree.bind(
            "<Configure>",
            lambda _event: self.after_idle(self._position_detail_number_widgets),
        )

        crm_controls = ttk.Frame(self.crm_tab)
        crm_controls.grid(row=0, column=0, sticky=tk.EW, pady=(0, 8))
        self.crm_search_keyword_var = tk.StringVar()
        ttk.Label(crm_controls, text="关键字").pack(side=tk.LEFT)
        crm_keyword_entry = ttk.Entry(
            crm_controls,
            textvariable=self.crm_search_keyword_var,
            width=36,
        )
        crm_keyword_entry.pack(side=tk.LEFT, padx=(6, 8))
        self.crm_search_button = ttk.Button(
            crm_controls,
            text="查询 CRM",
            command=lambda: self._query_crm_bugs(1),
            style="Accent.TButton",
        )
        self.crm_search_button.pack(side=tk.LEFT)
        ttk.Label(
            crm_controls,
            text="支持 Bug 编号或标题关键字，每页 50 条",
            style="Muted.TLabel",
        ).pack(side=tk.LEFT, padx=(10, 0))

        self.crm_next_button = ttk.Button(
            crm_controls,
            text="下一页",
            command=self._next_crm_page,
            state=tk.DISABLED,
            style="Secondary.TButton",
        )
        self.crm_next_button.pack(side=tk.RIGHT)
        self.crm_previous_button = ttk.Button(
            crm_controls,
            text="上一页",
            command=self._previous_crm_page,
            state=tk.DISABLED,
            style="Secondary.TButton",
        )
        self.crm_previous_button.pack(side=tk.RIGHT, padx=(0, 8))
        self.crm_page_status_var = tk.StringVar(value="尚未查询")
        ttk.Label(
            crm_controls,
            textvariable=self.crm_page_status_var,
        ).pack(side=tk.RIGHT, padx=(0, 12))

        crm_tree_frame = ttk.Frame(self.crm_tab)
        crm_tree_frame.grid(row=1, column=0, sticky=tk.NSEW)
        self.crm_tab.rowconfigure(1, weight=1)
        self.crm_tab.columnconfigure(0, weight=1)
        self.crm_tree = self._make_tree(
            crm_tree_frame,
            list(CRM_TABLE_COLUMNS),
        )
        self.crm_number_widgets: dict[
            str, tuple[tk.Frame, tk.Label, tk.Canvas]
        ] = {}
        self.crm_tree.bind("<Button-1>", self._handle_crm_tree_click)
        self.crm_tree.bind("<Motion>", self._update_crm_tree_cursor)
        self.crm_tree.bind(
            "<Leave>",
            lambda _event: self.crm_tree.configure(cursor=""),
        )
        self.crm_tree.bind(
            "<<TreeviewScroll>>",
            lambda _event: self.after_idle(self._position_crm_number_links),
        )
        self.crm_tree.bind(
            "<<TreeviewSelect>>",
            lambda _event: self.after_idle(self._position_crm_number_links),
        )
        self.crm_tree.bind(
            "<Configure>",
            lambda _event: self.after_idle(self._position_crm_number_links),
        )
        crm_keyword_entry.bind(
            "<Return>",
            lambda _event: self._query_crm_bugs(1),
        )
        self._build_workload_tab(self.workload_tab)
        self._select_main_tab("performance")

    def _select_main_tab(self, tab_name: str) -> None:
        tabs = {
            "performance": self.performance_detail_tab,
            "crm": self.crm_tab,
            "workload": self.workload_tab,
        }
        selected = tab_name if tab_name in tabs else "performance"
        self.notebook.select(tabs[selected])
        self.performance_nav_button.configure(
            style=(
                "NavActive.TButton"
                if selected == "performance"
                else "NavInactive.TButton"
            )
        )
        self.crm_nav_button.configure(
            style="NavActive.TButton" if selected == "crm" else "NavInactive.TButton"
        )
        self.workload_nav_button.configure(
            style=(
                "NavActive.TButton"
                if selected == "workload"
                else "NavInactive.TButton"
            )
        )
        if selected == "workload" and self.logged_in:
            if not self.workload_group_combo.cget("values"):
                self._load_workload_groups()

    def _make_tree(
        self,
        parent: ttk.Frame,
        columns: list[tuple[str, str, int]],
    ) -> ttk.Treeview:
        tree = ttk.Treeview(
            parent,
            columns=[item[0] for item in columns],
            show="headings",
            selectmode="browse",
        )
        vertical = ttk.Scrollbar(parent, orient=tk.VERTICAL, command=tree.yview)
        horizontal = ttk.Scrollbar(parent, orient=tk.HORIZONTAL, command=tree.xview)

        def update_vertical_scroll(first: str, last: str) -> None:
            vertical.set(first, last)
            tree.event_generate("<<TreeviewScroll>>", when="tail")

        def update_horizontal_scroll(first: str, last: str) -> None:
            horizontal.set(first, last)
            tree.event_generate("<<TreeviewScroll>>", when="tail")

        tree.configure(
            yscrollcommand=update_vertical_scroll,
            xscrollcommand=update_horizontal_scroll,
        )
        tree.tag_configure("evenrow", background=COLOR_SURFACE)
        tree.tag_configure("oddrow", background=COLOR_SURFACE_ALT)

        tree.grid(row=0, column=0, sticky=tk.NSEW)
        vertical.grid(row=0, column=1, sticky=tk.NS)
        horizontal.grid(row=1, column=0, sticky=tk.EW)
        parent.rowconfigure(0, weight=1)
        parent.columnconfigure(0, weight=1)

        for key, label, width in columns:
            tree.heading(key, text=label)
            tree.column(
                key,
                width=width,
                minwidth=min(width, 70),
                stretch=key in {"title", "module", "introducer"},
            )
        return tree

    def _run_worker(self, operation: str, function: Callable[[], Any]) -> None:
        if self.busy:
            return
        self._set_busy(True)

        def work() -> None:
            try:
                self.worker_messages.put((f"{operation}:success", function()))
            except Exception as exc:
                self.worker_messages.put((f"{operation}:error", exc))

        threading.Thread(target=work, daemon=True).start()

    def _poll_worker_messages(self) -> None:
        try:
            while True:
                operation, payload = self.worker_messages.get_nowait()
                self._handle_worker_message(operation, payload)
        except queue.Empty:
            pass
        self.after(100, self._poll_worker_messages)

    def _handle_worker_message(self, operation: str, payload: Any) -> None:
        self._set_busy(False)
        if operation == "login:success":
            self.logged_in = True
            self.login_status_label.configure(style="Connected.TLabel")
            performance_result, _crm_result = payload
            self.performance_display_name = performance_result.display_name
            self.introducer_var.set(performance_result.display_name)
            self.login_status_var.set(
                f"已登录：{performance_result.display_name}"
                + (
                    f"（{performance_result.role}，绩效 + CRM）"
                    if performance_result.role
                    else "（绩效 + CRM）"
                )
            )
            credentials_saved = save_credentials(
                CREDENTIALS_PATH,
                self.username_var.get().strip(),
                self.password_var.get(),
                self.crm_password_var.get(),
            )
            self.status_var.set(
                "登录成功，可以查询缺陷"
                if credentials_saved
                else "登录成功，但密码未能保存；本次运行中仍会保留"
            )
            self._save_settings()
            self._load_plan_versions()
            return
        if operation == "plans:success":
            self._apply_plan_versions(payload)
            return
        if operation == "workload_groups:success":
            self._apply_workload_groups(payload)
            return
        if operation == "workload_developers:success":
            developers = [ALL_GROUP_MEMBERS, *[item.name for item in payload]]
            self.workload_developer_combo.configure(values=developers)
            selected = self.workload_developer_var.get().strip()
            self.workload_developer_var.set(
                selected if selected in developers else ALL_GROUP_MEMBERS
            )
            self.status_var.set(
                f"已加载 {len(developers) - 1} 名当前分组前端成员"
            )
            self._save_settings()
            return
        if operation == "workload_preview:success":
            context, preview = payload
            self.workload_context = context
            self.workload_preview = preview
            developers = [ALL_GROUP_MEMBERS, *sorted(context.developers)]
            self.workload_developer_combo.configure(values=developers)
            if self.workload_developer_var.get() not in developers:
                self.workload_developer_var.set(ALL_GROUP_MEMBERS)
            self._render_workload_preview()
            self._save_settings()
            return
        if operation == "workload_submit:success":
            result: WorkloadSubmitResult = payload
            self.workload_context = None
            self.workload_preview = None
            self.workload_submit_button.configure(state=tk.DISABLED)
            self.status_var.set(
                f"绩效录入完成：提交 {result.submitted_count} 条，"
                f"回查确认 {result.verified_count} 条"
            )
            self.workload_summary_var.set(
                f"保存前 {result.before_count} 条，保存后 {result.after_count} 条；"
                "如需再次导入，请重新加载预览"
            )
            messagebox.showinfo(
                APP_TITLE,
                f"{result.message}\n\n"
                f"提交：{result.submitted_count} 条\n"
                f"回查确认：{result.verified_count} 条\n"
                f"分组记录数：{result.before_count} → {result.after_count}",
                parent=self,
            )
            return
        if operation == "query:success":
            self.all_bugs = payload
            self._apply_filter()
            self.status_var.set(
                f"查询成功：接口返回 {len(self.all_bugs)} 条，"
                f"当前筛选 {len(self.filtered_bugs)} 条"
            )
            self._save_settings()
            return
        if operation == "crm_search:success":
            self.crm_bug_page = payload
            self._render_crm_bugs()
            self._update_crm_paging_buttons()
            self._select_main_tab("crm")
            self.status_var.set(
                f"CRM 查询成功：共 {payload.total} 条，"
                f"当前第 {payload.page_num} 页显示 {len(payload.items)} 条"
            )
            return
        if operation == "detail:success":
            bug, detail, image_data = payload
            self.status_var.set(
                f"已读取 {detail.bug_number or text(bug.get('id'))} 的详情"
            )
            BugDetailDialog(self, bug, detail, image_data)
            return
        if operation == "md_export:success":
            output_path, bug_count, unavailable_count = payload
            self.status_var.set(f"已导出 Bug 回溯：{output_path}")
            unavailable_hint = (
                f"\n其中 {unavailable_count} 条 CRM 详情读取失败，"
                "文档中已保留缺陷基础信息。"
                if unavailable_count
                else ""
            )
            messagebox.showinfo(
                APP_TITLE,
                f"Bug 回溯导出完成，共 {bug_count} 条。\n{output_path}"
                f"{unavailable_hint}",
                parent=self,
            )
            return

        if operation.endswith(":error"):
            if isinstance(
                payload,
                (AuthenticationError, CrmAuthenticationError),
            ):
                self.logged_in = False
                self.login_status_var.set("未登录")
                self.login_status_label.configure(style="LoginStatus.TLabel")
            self.status_var.set(str(payload))
            messagebox.showerror(APP_TITLE, str(payload), parent=self)

    def _set_busy(self, busy: bool) -> None:
        self.busy = busy
        state = tk.DISABLED if busy else tk.NORMAL
        for button in [
            self.login_button,
            self.query_button,
            self.filter_button,
            self.clear_button,
            self.export_button,
            self.export_md_button,
            self.crm_search_button,
            self.crm_previous_button,
            self.crm_next_button,
            self.workload_browse_button,
            self.workload_preview_button,
            self.workload_clear_button,
            self.workload_submit_button,
        ]:
            button.configure(state=state)
        if busy:
            self.progress.start(10)
            self._show_loading()
        else:
            self.progress.stop()
            self._hide_loading()
            self._update_crm_paging_buttons()
            self.workload_submit_button.configure(
                state=(
                    tk.NORMAL
                    if self.workload_context
                    and self.workload_preview
                    and self.workload_preview.submittable_rows
                    else tk.DISABLED
                )
            )

    def _login(self) -> None:
        username = self.username_var.get()
        password = self.password_var.get()
        crm_password = self.crm_password_var.get()
        self.status_var.set("正在登录绩效系统和 CRM…")

        def login_both() -> tuple[Any, Any]:
            performance_result = self.client.login(username, password)
            crm_result = self.crm_client.login(username, crm_password)
            return performance_result, crm_result

        self._run_worker("login", login_both)

    def _query(self) -> None:
        if not self.logged_in:
            messagebox.showinfo(APP_TITLE, "请先登录", parent=self)
            return
        plan_version = self.plan_var.get()
        self.status_var.set(f"正在查询计划版本 {plan_version.strip()}…")
        self._run_worker(
            "query",
            lambda: self.client.get_bugs_by_plan(plan_version),
        )

    def _load_plan_versions(self) -> None:
        current_year = datetime.now().year
        self.status_var.set(f"正在加载 {current_year} 年计划版本…")
        self._run_worker(
            "plans",
            lambda: self.client.get_plans_by_year(current_year),
        )

    def _apply_plan_versions(self, plans: list[PlanVersion]) -> None:
        self.plan_versions = plans
        values = [plan.plan_version for plan in plans]
        self.plan_combo.configure(values=values)
        self.workload_plan_combo.configure(values=values)
        if not plans:
            self.status_var.set("当前年份没有可选的计划版本")
            return
        now = datetime.now()
        selected = next(
            (
                plan.plan_version
                for plan in plans
                if plan.year == now.year and plan.month == now.month
            ),
            "",
        )
        if not selected:
            saved = self.settings.get("plan_version", "")
            selected = saved if saved in values else values[0]
        self.plan_var.set(selected)
        saved_workload_plan = self.workload_plan_var.get().strip()
        self.workload_plan_var.set(
            saved_workload_plan if saved_workload_plan in values else selected
        )
        self.status_var.set(
            f"已加载 {now.year} 年 {len(plans)} 个计划版本，"
            f"当前选择 {selected}"
        )
        self._save_settings()

    def _load_workload_groups(self) -> None:
        if not self.logged_in:
            return
        self.status_var.set("正在读取可录入的绩效分组…")
        self._run_worker("workload_groups", self.workload_api.get_groups)

    def _apply_workload_groups(self, groups: list[Any]) -> None:
        values = [group.name for group in groups]
        self.workload_group_combo.configure(values=values)
        if not values:
            self.workload_group_var.set("")
            self.status_var.set("没有读取到可录入的开发分组")
            return
        saved = self.workload_group_var.get().strip()
        selected = (
            saved
            if saved in values
            else DEFAULT_WORKLOAD_GROUP
            if DEFAULT_WORKLOAD_GROUP in values
            else values[0]
        )
        self.workload_group_var.set(selected)
        self.status_var.set(f"已加载 {len(values)} 个绩效录入分组")
        self._save_settings()
        self._load_workload_developers()

    def _load_workload_developers(self) -> None:
        if not self.logged_in:
            return
        module = self.workload_group_var.get().strip()
        if not module:
            return
        self.status_var.set(f"正在读取 {module} 的开发资源…")
        self._run_worker(
            "workload_developers",
            lambda: self.workload_api.get_developers(module),
        )

    def _choose_workload_file(self) -> None:
        initial_path = self.workload_file_var.get().strip()
        initial_directory = (
            str(Path(initial_path).parent)
            if initial_path and Path(initial_path).parent.exists()
            else str(APP_DIR.parent)
        )
        selected = filedialog.askopenfilename(
            parent=self,
            title="选择工作量 Excel",
            initialdir=initial_directory,
            filetypes=[("Excel 工作簿", "*.xlsx")],
        )
        if not selected:
            return
        self.workload_file_var.set(selected)
        self._clear_workload_preview()
        self.status_var.set(f"已选择 Excel：{Path(selected).name}")
        self._save_settings()

    def _clear_workload_preview(self) -> None:
        self.workload_context = None
        self.workload_preview = None
        if hasattr(self, "workload_tree"):
            self.workload_tree.delete(*self.workload_tree.get_children())
        if hasattr(self, "workload_summary_var"):
            self.workload_summary_var.set("请选择 Excel，加载后先预览校验结果")
        if hasattr(self, "workload_submit_button"):
            self.workload_submit_button.configure(state=tk.DISABLED)

    def _workload_group_changed(self, _event: tk.Event | None = None) -> None:
        self.workload_developer_var.set(ALL_GROUP_MEMBERS)
        self.workload_developer_combo.configure(values=[ALL_GROUP_MEMBERS])
        self._clear_workload_preview()
        self._save_settings()
        self._load_workload_developers()

    def _preview_workload(self) -> None:
        if not self.logged_in:
            messagebox.showinfo(APP_TITLE, "请先登录", parent=self)
            return
        plan_version = self.workload_plan_var.get().strip()
        module = self.workload_group_var.get().strip()
        excel_path = self.workload_file_var.get().strip()
        developer_filter = self.workload_developer_var.get().strip() or ALL_GROUP_MEMBERS
        if not plan_version:
            messagebox.showinfo(APP_TITLE, "请选择计划版本", parent=self)
            return
        if not module:
            messagebox.showinfo(APP_TITLE, "请选择录入分组", parent=self)
            return
        if not excel_path:
            messagebox.showinfo(APP_TITLE, "请先选择 Excel 文件", parent=self)
            return

        self._clear_workload_preview()
        self.status_var.set(
            f"正在校验 {Path(excel_path).name}，目标 {plan_version} / {module}…"
        )

        def build_preview() -> tuple[WorkloadContext, WorkloadPreview]:
            context = self.workload_api.get_context(plan_version, module)
            items = read_workload_excel(excel_path, plan_version)
            preview = build_workload_preview(items, context, developer_filter)
            return context, preview

        self._run_worker("workload_preview", build_preview)

    def _render_workload_preview(self) -> None:
        self.workload_tree.delete(*self.workload_tree.get_children())
        preview = self.workload_preview
        if preview is None:
            return
        tag_by_status = {
            "可提交": "valid",
            "提醒": "warning",
            "错误": "error",
            "已存在": "duplicate",
        }
        for row in preview.rows:
            source = row.source
            self.workload_tree.insert(
                "",
                tk.END,
                values=(
                    source.excel_row,
                    row.status,
                    source.developer_name,
                    source.require_no,
                    source.task_name,
                    source.plan_start_date,
                    source.plan_finish_date,
                    f"{source.requested_days * 8:g}",
                    f"{row.computed_hours:g}",
                    row.message,
                ),
                tags=(tag_by_status.get(row.status, ""),),
            )
        ignored_hint = (
            "；非本组责任人：" + "、".join(preview.ignored_developers)
            if preview.ignored_developers
            else ""
        )
        self.workload_summary_var.set(
            f"Excel 有效数据 {preview.source_row_count} 行，匹配 {len(preview.rows)} 行，"
            f"可提交 {len(preview.submittable_rows)} 行（提醒 {preview.warning_count}），"
            f"错误 {preview.error_count} 行，已存在 {preview.duplicate_count} 行，"
            f"跳过 {preview.skipped_row_count} 行；"
            f"Excel {preview.requested_hours:g}h / 计算 {preview.computed_hours:g}h"
            f"{ignored_hint}"
        )
        self.status_var.set(
            f"预览完成：{len(preview.submittable_rows)} 条可以提交，"
            f"{preview.error_count} 条需要修正"
        )
        self.workload_submit_button.configure(
            state=tk.NORMAL if preview.submittable_rows else tk.DISABLED
        )

    def _submit_workload(self) -> None:
        context = self.workload_context
        preview = self.workload_preview
        if context is None or preview is None or not preview.submittable_rows:
            messagebox.showinfo(APP_TITLE, "请先加载并检查导入预览", parent=self)
            return
        count = len(preview.submittable_rows)
        confirmed = messagebox.askyesno(
            APP_TITLE,
            "即将向绩效系统写入工作量：\n\n"
            f"计划版本：{context.plan_version}\n"
            f"录入分组：{context.module}\n"
            f"新增记录：{count} 条\n"
            f"计算工时：{preview.computed_hours:g} 小时\n\n"
            "系统会一次批量保存，并在保存后自动回查。确认提交吗？",
            icon="warning",
            parent=self,
        )
        if not confirmed:
            return
        rows = list(preview.submittable_rows)
        self.status_var.set(f"正在批量保存 {count} 条绩效工作量…")
        self._run_worker(
            "workload_submit",
            lambda: self.workload_api.submit(context, rows),
        )

    def _query_crm_bugs(
        self,
        page_num: int,
        *,
        use_active_keyword: bool = False,
    ) -> None:
        if not self.logged_in:
            messagebox.showinfo(APP_TITLE, "请先登录", parent=self)
            return
        keyword = (
            self.crm_bug_page.keyword
            if use_active_keyword and self.crm_bug_page
            else self.crm_search_keyword_var.get()
        )
        self.status_var.set(
            f"正在查询 CRM 缺陷第 {max(1, page_num)} 页…"
        )
        self._run_worker(
            "crm_search",
            lambda: self.crm_client.search_bugs(keyword, page_num),
        )

    def _previous_crm_page(self) -> None:
        if self.crm_bug_page and self.crm_bug_page.page_num > 1:
            self._query_crm_bugs(
                self.crm_bug_page.page_num - 1,
                use_active_keyword=True,
            )

    def _next_crm_page(self) -> None:
        if (
            self.crm_bug_page
            and self.crm_bug_page.page_num < self.crm_bug_page.pages
        ):
            self._query_crm_bugs(
                self.crm_bug_page.page_num + 1,
                use_active_keyword=True,
            )

    def _update_crm_paging_buttons(self) -> None:
        if self.busy or not self.crm_bug_page:
            self.crm_previous_button.configure(state=tk.DISABLED)
            self.crm_next_button.configure(state=tk.DISABLED)
            return
        self.crm_previous_button.configure(
            state=(
                tk.NORMAL
                if self.crm_bug_page.page_num > 1
                else tk.DISABLED
            )
        )
        self.crm_next_button.configure(
            state=(
                tk.NORMAL
                if self.crm_bug_page.page_num < self.crm_bug_page.pages
                else tk.DISABLED
            )
        )

    def _render_crm_bugs(self) -> None:
        for frame, _label, _icon in self.crm_number_widgets.values():
            frame.destroy()
        self.crm_number_widgets.clear()
        self.crm_tree.delete(*self.crm_tree.get_children())
        if not self.crm_bug_page:
            self.crm_page_status_var.set("尚未查询")
            return
        for index, bug in enumerate(self.crm_bug_page.items):
            values = [
                bug.get(key, "")
                for key, _label, _width in CRM_TABLE_COLUMNS
            ]
            self.crm_tree.insert(
                "",
                tk.END,
                iid=str(index),
                values=values,
                tags=("evenrow" if index % 2 == 0 else "oddrow",),
            )
        shown_page = self.crm_bug_page.page_num if self.crm_bug_page.total else 0
        self.crm_page_status_var.set(
            f"第 {shown_page}/{self.crm_bug_page.pages} 页，"
            f"共 {self.crm_bug_page.total} 条"
        )
        self.after_idle(self._position_crm_number_links)

    def _crm_tree_column_key(self, x_position: int) -> str:
        column_id = self.crm_tree.identify_column(x_position)
        if not column_id.startswith("#"):
            return ""
        try:
            column_index = int(column_id[1:]) - 1
            return str(self.crm_tree["columns"][column_index])
        except (ValueError, IndexError):
            return ""

    def _update_crm_tree_cursor(self, event: tk.Event) -> None:
        is_link_cell = (
            self.crm_tree.identify_region(event.x, event.y) == "cell"
            and self._crm_tree_column_key(event.x) == "objectNumber"
        )
        self.crm_tree.configure(cursor="hand2" if is_link_cell else "")

    def _handle_crm_tree_click(self, event: tk.Event) -> str | None:
        if (
            self.crm_tree.identify_region(event.x, event.y) != "cell"
            or not self.crm_bug_page
        ):
            return None
        column_key = self._crm_tree_column_key(event.x)
        if column_key != "objectNumber":
            return None
        item_id = self.crm_tree.identify_row(event.y)
        if not item_id:
            return None
        return self._open_crm_detail_from_link(item_id)

    def _copy_bug_number(self, bug_number: str) -> None:
        if not bug_number:
            return
        self.clipboard_clear()
        self.clipboard_append(bug_number)
        self.update_idletasks()
        self.status_var.set(f"复制成功：{bug_number}")
        toast = tk.Label(
            self,
            text="✓  复制成功",
            background=COLOR_SUCCESS,
            foreground="#FFFFFF",
            padx=18,
            pady=8,
            borderwidth=0,
            font=("Microsoft YaHei UI", 10, "bold"),
        )
        toast.place(relx=0.5, rely=0.94, anchor=tk.S)
        toast.lift()

        def hide_toast() -> None:
            if toast.winfo_exists():
                toast.destroy()

        self.after(1500, hide_toast)

    @staticmethod
    def _make_copy_icon(
        parent: tk.Misc,
        background: str,
        command: Callable[[], None],
    ) -> tk.Canvas:
        icon = tk.Canvas(
            parent,
            width=18,
            height=18,
            background=background,
            highlightthickness=0,
            borderwidth=0,
            cursor="hand2",
        )
        icon.create_rectangle(
            6,
            3,
            14,
            11,
            outline=COLOR_ACCENT,
            width=1,
            tags=("copy_icon",),
        )
        icon.create_rectangle(
            3,
            6,
            11,
            14,
            outline=COLOR_ACCENT,
            width=1,
            tags=("copy_icon",),
        )

        def copy_clicked(_event: tk.Event) -> str:
            command()
            return "break"

        icon.bind("<Button-1>", copy_clicked)
        return icon

    @staticmethod
    def _configure_copy_icon(
        icon: tk.Canvas,
        background: str,
        foreground: str,
    ) -> None:
        icon.configure(background=background)
        icon.itemconfigure("copy_icon", outline=foreground)

    def _position_crm_number_links(self) -> None:
        if not self.crm_tree.winfo_exists():
            return
        selected = set(self.crm_tree.selection())
        for item_id in self.crm_tree.get_children():
            bounds = self.crm_tree.bbox(item_id, "objectNumber")
            widgets = self.crm_number_widgets.get(item_id)
            if not bounds:
                if widgets:
                    widgets[0].place_forget()
                continue
            try:
                bug = self.crm_bug_page.items[int(item_id)] if self.crm_bug_page else {}
            except (ValueError, IndexError):
                bug = {}
            bug_number = text(bug.get("objectNumber"))
            if not widgets:
                frame = tk.Frame(self.crm_tree, borderwidth=0)
                label = tk.Label(
                    frame,
                    text=bug_number,
                    anchor=tk.W,
                    padx=0,
                    borderwidth=0,
                    cursor="hand2",
                    font=("Microsoft YaHei UI", 9),
                )
                label.pack(side=tk.LEFT, padx=(4, 0))
                label.bind(
                    "<Button-1>",
                    lambda _event, iid=item_id: self._open_crm_detail_from_link(iid),
                )
                icon = self._make_copy_icon(
                    frame,
                    COLOR_SURFACE,
                    lambda number=bug_number: self._copy_bug_number(number),
                )
                icon.pack(side=tk.LEFT, padx=(3, 0))
                widgets = (frame, label, icon)
                self.crm_number_widgets[item_id] = widgets
            frame, label, icon = widgets
            is_selected = item_id in selected
            row_index = int(item_id)
            background = (
                COLOR_ACCENT
                if is_selected
                else COLOR_SURFACE if row_index % 2 == 0 else COLOR_SURFACE_ALT
            )
            foreground = "#FFFFFF" if is_selected else COLOR_ACCENT
            frame.configure(background=background)
            label.configure(
                text=bug_number,
                background=background,
                foreground=foreground,
            )
            self._configure_copy_icon(icon, background, foreground)
            x, y, width, height = bounds
            frame.place(
                x=x + 1,
                y=y + 1,
                width=max(1, width - 2),
                height=max(1, height - 2),
            )
            frame.lift()

    def _open_crm_detail_from_link(self, item_id: str) -> str:
        self.crm_tree.selection_set(item_id)
        self.crm_tree.focus(item_id)
        self._show_crm_bug_detail(item_id)
        return "break"

    def _apply_filter(self) -> None:
        self.filtered_bugs = filter_bugs(
            self.all_bugs,
            introducer=self.introducer_var.get(),
            keyword=self.keyword_var.get(),
        )
        self._render_details()
        self._render_metrics()
        self.status_var.set(
            f"接口数据 {len(self.all_bugs)} 条，当前筛选 {len(self.filtered_bugs)} 条"
        )

    def _clear_filter(self) -> None:
        self.introducer_var.set("")
        self.keyword_var.set("")
        self._apply_filter()

    def _render_details(self) -> None:
        for frame, _label, _icon in self.detail_number_widgets.values():
            frame.destroy()
        self.detail_number_widgets.clear()
        self.detail_tree.delete(*self.detail_tree.get_children())
        for index, bug in enumerate(self.filtered_bugs):
            values = [text(bug.get(key)) for key, _label, _width in DETAIL_COLUMNS]
            self.detail_tree.insert(
                "",
                tk.END,
                iid=str(index),
                values=values,
                tags=("evenrow" if index % 2 == 0 else "oddrow",),
            )
        self.after_idle(self._position_detail_number_widgets)

    def _position_detail_number_widgets(self) -> None:
        if not self.detail_tree.winfo_exists():
            return
        selected = set(self.detail_tree.selection())
        for item_id in self.detail_tree.get_children():
            bounds = self.detail_tree.bbox(item_id, "id")
            widgets = self.detail_number_widgets.get(item_id)
            if not bounds:
                if widgets:
                    widgets[0].place_forget()
                continue
            try:
                bug = self.filtered_bugs[int(item_id)]
            except (ValueError, IndexError):
                bug = {}
            bug_number = text(bug.get("id"))
            if not widgets:
                frame = tk.Frame(self.detail_tree, borderwidth=0)
                label = tk.Label(
                    frame,
                    text=bug_number,
                    anchor=tk.W,
                    padx=0,
                    borderwidth=0,
                    font=("Microsoft YaHei UI", 9),
                )
                label.pack(side=tk.LEFT, padx=(4, 0))
                label.bind(
                    "<Button-1>",
                    lambda _event, iid=item_id: self._open_detail_from_link(iid),
                )
                icon = self._make_copy_icon(
                    frame,
                    COLOR_SURFACE,
                    lambda number=bug_number: self._copy_bug_number(number),
                )
                icon.pack(side=tk.LEFT, padx=(3, 0))
                widgets = (frame, label, icon)
                self.detail_number_widgets[item_id] = widgets
            frame, label, icon = widgets
            is_selected = item_id in selected
            row_index = int(item_id)
            background = (
                COLOR_ACCENT
                if is_selected
                else COLOR_SURFACE if row_index % 2 == 0 else COLOR_SURFACE_ALT
            )
            foreground = "#FFFFFF" if is_selected else COLOR_ACCENT
            frame.configure(background=background)
            label.configure(
                text=bug_number,
                background=background,
                foreground=foreground,
            )
            self._configure_copy_icon(
                icon,
                background,
                "#FFFFFF" if is_selected else COLOR_ACCENT,
            )
            x, y, width, height = bounds
            frame.place(
                x=x + 1,
                y=y + 1,
                width=max(1, width - 2),
                height=max(1, height - 2),
            )
            frame.lift()

    def _open_detail_from_link(self, item_id: str) -> str:
        self.detail_tree.selection_set(item_id)
        self.detail_tree.focus(item_id)
        return self._show_bug_detail_by_item(item_id)

    def _render_metrics(self) -> None:
        metrics = summary_metrics(self.filtered_bugs)
        for key, variable in self.metric_vars.items():
            variable.set(str(metrics[key]))

    def _show_bug_detail_by_item(self, item_id: str) -> str:
        try:
            index = int(item_id)
        except ValueError:
            return "break"
        if not 0 <= index < len(self.filtered_bugs):
            return "break"
        bug = self.filtered_bugs[index]
        url = text(bug.get("bugUrl"))
        if not url:
            messagebox.showerror(APP_TITLE, "该缺陷没有 CRM 链接", parent=self)
            return "break"
        self.status_var.set(f"正在读取 {text(bug.get('id'))} 的详情…")

        def load_detail() -> tuple[dict[str, Any], BugDetail, dict[str, bytes]]:
            detail = self.crm_client.get_bug_detail(url)
            sources: list[str] = []
            for field in detail.fields:
                if field.is_rich_text:
                    sources.extend(extract_image_sources(field.value))
            image_data = self.crm_client.download_rich_images(sources)
            return bug, detail, image_data

        self._run_worker(
            "detail",
            load_detail,
        )
        return "break"

    def _show_crm_bug_detail(self, item_id: str) -> None:
        if not self.crm_bug_page:
            return
        try:
            index = int(item_id)
        except ValueError:
            return
        if not 0 <= index < len(self.crm_bug_page.items):
            return
        crm_bug = self.crm_bug_page.items[index]
        object_oid = crm_bug.get("_oid", "")
        object_otype = crm_bug.get("_otype", "")
        if not object_oid or not object_otype:
            messagebox.showerror(
                APP_TITLE,
                "该 CRM 缺陷缺少对象标识，无法读取详情",
                parent=self,
            )
            return
        bug = {
            "id": crm_bug.get("objectNumber", ""),
            "title": crm_bug.get("title", ""),
            "bugUrl": build_crm_bug_url(
                self.crm_client.base_url,
                object_oid,
                object_otype,
                text(crm_bug.get("objectNumber")),
            ),
        }
        self.status_var.set(
            f"正在读取 {crm_bug.get('objectNumber', '')} 的详情…"
        )

        def load_detail() -> tuple[dict[str, Any], BugDetail, dict[str, bytes]]:
            detail = self.crm_client.get_bug_detail_by_identity(
                object_oid,
                object_otype,
            )
            sources: list[str] = []
            for field in detail.fields:
                if field.is_rich_text:
                    sources.extend(extract_image_sources(field.value))
            image_data = self.crm_client.download_rich_images(sources)
            return bug, detail, image_data

        self._run_worker("detail", load_detail)

    def _export_excel(self) -> None:
        if not self.filtered_bugs:
            messagebox.showinfo(APP_TITLE, "当前没有可导出的缺陷数据", parent=self)
            return
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError:
            messagebox.showerror(
                APP_TITLE,
                "缺少 openpyxl，请运行：pip install openpyxl",
                parent=self,
            )
            return

        plan = self.plan_var.get().strip() or "未命名版本"
        suggested = f"缺陷统计_{plan}_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
        path = filedialog.asksaveasfilename(
            parent=self,
            title="导出缺陷统计",
            defaultextension=".xlsx",
            initialfile=suggested,
            filetypes=[("Excel 工作簿", "*.xlsx")],
        )
        if not path:
            return

        workbook = Workbook()
        summary_sheet = workbook.active
        summary_sheet.title = "统计汇总"
        detail_sheet = workbook.create_sheet("缺陷明细")
        header_fill = PatternFill("solid", fgColor="D9EAF7")
        header_font = Font(bold=True)

        summary_headers = [label for _key, label, _width in SUMMARY_COLUMNS]
        summary_sheet.append(summary_headers)
        for row in introducer_summary(self.all_bugs):
            summary_sheet.append([row[key] for key, _label, _width in SUMMARY_COLUMNS])

        detail_headers = [label for _key, label, _width in DETAIL_COLUMNS] + ["缺陷链接"]
        detail_sheet.append(detail_headers)
        for bug in self.filtered_bugs:
            detail_sheet.append(
                [text(bug.get(key)) for key, _label, _width in DETAIL_COLUMNS]
                + [text(bug.get("bugUrl"))]
            )

        for sheet in (summary_sheet, detail_sheet):
            sheet.freeze_panes = "A2"
            sheet.auto_filter.ref = sheet.dimensions
            for cell in sheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

        for index, (_key, _label, width) in enumerate(SUMMARY_COLUMNS, 1):
            summary_sheet.column_dimensions[get_column_letter(index)].width = max(
                10, width / 7
            )
        for index, (_key, _label, width) in enumerate(DETAIL_COLUMNS, 1):
            detail_sheet.column_dimensions[get_column_letter(index)].width = max(
                10, min(width / 7, 70)
            )
        detail_sheet.column_dimensions[
            get_column_letter(len(DETAIL_COLUMNS) + 1)
        ].width = 60

        try:
            workbook.save(path)
        except OSError as exc:
            messagebox.showerror(APP_TITLE, f"导出失败：{exc}", parent=self)
            return
        self.status_var.set(f"已导出：{path}")
        messagebox.showinfo(APP_TITLE, "Excel 导出完成", parent=self)

    def _export_retrospective_md(self) -> None:
        if not self.filtered_bugs:
            messagebox.showinfo(
                APP_TITLE,
                "当前没有可生成回溯的缺陷数据",
                parent=self,
            )
            return
        if not self.logged_in:
            messagebox.showinfo(APP_TITLE, "请先登录", parent=self)
            return

        plan_version = self.plan_var.get().strip()
        introducer = self.introducer_var.get().strip()
        suggested = retrospective_filename(plan_version, introducer)
        preferred_dir = (
            Path.home()
            / "Desktop"
            / "需求转测试+设计文档"
            / "7月份需求"
        )
        dialog_options: dict[str, Any] = {
            "parent": self,
            "title": "导出 Bug 回溯",
            "defaultextension": ".md",
            "initialfile": suggested,
            "filetypes": [("Markdown 文档", "*.md")],
        }
        if preferred_dir.exists():
            dialog_options["initialdir"] = str(preferred_dir)
        output_path = filedialog.asksaveasfilename(**dialog_options)
        if not output_path:
            return

        bugs = list(self.filtered_bugs)
        self.status_var.set(
            f"正在读取 {len(bugs)} 条 CRM 详情并生成 Bug 回溯…"
        )

        def build_export() -> tuple[str, int, int]:
            details: dict[str, BugDetail | None] = {}
            unavailable_count = 0
            for bug in bugs:
                bug_number = text(bug.get("id"))
                bug_url = text(bug.get("bugUrl"))
                if not bug_url:
                    details[bug_number] = None
                    unavailable_count += 1
                    continue
                try:
                    details[bug_number] = self.crm_client.get_bug_detail(bug_url)
                except CrmAuthenticationError:
                    raise
                except CrmClientError:
                    details[bug_number] = None
                    unavailable_count += 1
            markdown = build_retrospective_markdown(
                plan_version,
                introducer,
                bugs,
                details,
            )
            Path(output_path).write_text(markdown, encoding="utf-8")
            return output_path, len(bugs), unavailable_count

        self._run_worker("md_export", build_export)

    def _on_close(self) -> None:
        self._save_settings()
        self.client.close()
        self.crm_client.close()
        self.destroy()


if __name__ == "__main__":
    if len(sys.argv) >= 4 and sys.argv[1] == "--diagnose-workload-excel":
        diagnostic_output = Path(sys.argv[3])
        try:
            diagnostic_items = read_workload_excel(sys.argv[2], "2026-0830")
            diagnostic_result = {
                "success": True,
                "rowCount": len(diagnostic_items),
            }
        except Exception as exc:
            diagnostic_result = {"success": False, "message": str(exc)}
        diagnostic_output.write_text(
            json.dumps(diagnostic_result, ensure_ascii=False),
            encoding="utf-8",
        )
    else:
        app = BugStatisticsApp()
        app.mainloop()
