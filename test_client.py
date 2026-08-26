"""不连接线上服务的核心逻辑测试。"""

from pathlib import Path
from tempfile import TemporaryDirectory
from unittest.mock import patch

from tkinter import ttk
from openpyxl import Workbook, load_workbook

from analytics import filter_bugs, introducer_name, introducer_summary, summary_metrics
from credential_store import load_credentials, save_credentials
from crm_client import (
    CRM_SEARCH_PAGE_SIZE,
    BugDetail,
    BugDetailField,
    build_crm_bug_url,
    build_crm_search_payload,
    extract_detail_fields,
    extract_crm_object_identity,
    encrypt_crm_password,
    normalize_crm_bug,
)
from detail_utils import (
    extract_image_sources,
    rich_content_segments,
    rich_text_to_plain,
)
from retrospective import (
    build_retrospective_markdown,
    retrospective_filename,
)
from workload_client import (
    ALL_GROUP_MEMBERS,
    SERVER_WORKLOAD_DEFINITION,
    UI_WORKLOAD_DEFINITION,
    Developer,
    ExcelWorkItem,
    WorkloadApiClient,
    WorkloadContext,
    allocate_ui_workload,
    allocate_server_workload,
    build_workload_preview,
    calculate_workload_hours,
    read_workload_excel,
)


SAMPLE = [
    {
        "id": "Bug001",
        "createdByName": "危国",
        "createdBy": "weig",
        "severity": "2",
        "status": "关闭",
        "title": "导出失败",
        "module": "公共模块",
    },
    {
        "id": "Bug002",
        "createdByName": "危国",
        "severity": "3",
        "status": "处理中",
        "title": "样式异常",
        "module": "表单管理",
    },
    {
        "id": "Bug003",
        "createdByName": "张三",
        "severity": "2",
        "status": "关闭",
        "title": "查询失败",
        "module": "公共模块",
    },
]


def test_analytics() -> None:
    assert introducer_name(SAMPLE[0]) == "危国"
    filtered = filter_bugs(SAMPLE, introducer="危国")
    assert [item["id"] for item in filtered] == ["Bug001", "Bug002"]
    assert len(filter_bugs(SAMPLE, keyword="样式")) == 1

    metrics = summary_metrics(filtered)
    assert metrics["total"] == 2
    assert metrics["severity_2"] == 1
    assert metrics["severity_3"] == 1
    assert metrics["closed"] == 1
    assert metrics["open"] == 1

    rows = introducer_summary(SAMPLE)
    assert rows[0]["introducer"] == "危国"
    assert rows[0]["total"] == 2


def test_crm_helpers() -> None:
    assert (
        encrypt_crm_password("example-password")
        == "UU9iMVo0MjFiOEFpREtXWDY0YUE4YU8yOXRiWk9Rb1N2azNCTmVNS3I3QT0="
    )
    url = (
        'https://crm.example/?objOrginalParam=%7B%22oid%22%3A%22123%22%2C'
        '%22otype%22%3A%22ty.inteplm.ipd.CTyDefect%22%7D'
    )
    assert extract_crm_object_identity(url) == (
        "123",
        "ty.inteplm.ipd.CTyDefect",
    )
    assert build_crm_bug_url(
        "https://crm.example/",
        "123",
        "ty.inteplm.ipd.CTyDefect",
        "Bug 001",
    ) == (
        "https://crm.example/#/homePage/defectObjForm?"
        "otype=ty.inteplm.ipd.CTyDefect&oid=123&label=Bug+001"
    )
    assert rich_text_to_plain("<p>原因：时序问题</p><p><img src='x'></p>") == (
        "原因：时序问题\n[图片]"
    )
    rich = "<p>验证截图：</p><p><img src='/rest/image?a=1&amp;b=2'></p>"
    assert extract_image_sources(rich) == ["/rest/image?a=1&b=2"]
    assert rich_content_segments(rich) == [
        ("text", "验证截图：\n"),
        ("image", "/rest/image?a=1&b=2"),
    ]


def test_credential_store() -> None:
    with TemporaryDirectory() as directory:
        path = Path(directory) / "credentials.dat"
        assert save_credentials(path, "tester", "performance-pass", "crm-pass")
        encrypted = path.read_bytes()
        assert b"performance-pass" not in encrypted
        assert b"crm-pass" not in encrypted
        assert load_credentials(path) == {
            "username": "tester",
            "password": "performance-pass",
            "crm_password": "crm-pass",
        }


def test_all_detail_fields() -> None:
    form_data = {
        "name": "示例缺陷",
        "owner": {"name": "T0001", "fullname": "测试用户"},
        "ibaAttribute": {
            "phase": ["SIT"],
            "IsAutoDetected": ["0"],
            "DTS_TestResult": ["<p>测试通过</p>"],
        },
    }
    template = {
        "widgetList": [
            {
                "type": "input",
                "options": {"realProps": "name", "label": "标题"},
            },
            {
                "type": "select-user",
                "options": {"realProps": "owner", "label": "负责人"},
            },
            {
                "type": "radio",
                "options": {
                    "realProps": "phase",
                    "label": "阶段",
                    "optionItems": [{"value": "SIT", "label": "SIT"}],
                },
            },
            {
                "type": "radio",
                "options": {
                    "realProps": "IsAutoDetected",
                    "label": "是否自动化发现",
                    "optionItems": [
                        {"value": "0", "label": "否"},
                        {"value": "1", "label": "是"},
                    ],
                },
            },
            {
                "type": "date",
                "options": {"realProps": "plandate", "label": "期望完成日期"},
            },
            {
                "type": "rich-editor",
                "options": {"realProps": "DTS_TestResult", "label": "测试结果"},
            },
        ]
    }
    fields = extract_detail_fields(form_data, template)
    assert [field.label for field in fields] == [
        "标题",
        "负责人",
        "阶段",
        "是否自动化发现",
        "期望完成日期",
        "测试结果",
    ]
    assert [field.value for field in fields[:-1]] == [
        "示例缺陷",
        "测试用户",
        "SIT",
        "否",
        "（空）",
    ]
    assert fields[-1].is_rich_text
    assert fields[-1].value == "<p>测试通过</p>"


def test_crm_search_helpers() -> None:
    count_payload = build_crm_search_payload(
        "Bug20260723003",
        1,
        need_count=True,
    )
    assert count_payload["pageNum"] == 0
    assert count_payload["pageSize"] == CRM_SEARCH_PAGE_SIZE == 50
    assert count_payload["needCount"] is True
    assert "isSetColumn" not in count_payload

    list_payload = build_crm_search_payload(
        "Bug20260723003",
        2,
        need_count=False,
    )
    assert list_payload["pageNum"] == 2
    assert list_payload["pageSize"] == 50
    assert list_payload["needCount"] is False
    assert list_payload["isSetColumn"] is False

    normalized = normalize_crm_bug(
        {
            "oid": "123",
            "otype": "ty.inteplm.ipd.CTyDefect",
            "objectNumber": "Bug001",
            "name": "回退标题",
            "lifecyclestagekeyName": "分析中",
            "creatorDisplayName": "测试用户(T0001)",
            "ibaAttrMapForSearch": {
                "827308980874936320": "2.0730",
                "773662542538473472": "公共模块 | PubSVC",
            },
            "extAttrMapForSearch": {
                "title": "缺陷标题",
                "defectCategory": "代码错误",
                "severity": "严重",
                "priorityName": "普通",
                "leader": "负责人(T0002)",
            },
        }
    )
    assert normalized["objectNumber"] == "Bug001"
    assert normalized["title"] == "缺陷标题"
    assert normalized["827308980874936320"] == "2.0730"
    assert normalized["773662542538473472"] == "公共模块 | PubSVC"
    assert normalized["_oid"] == "123"


def test_retrospective_markdown() -> None:
    fields = (
        BugDetailField(
            "description",
            "描述",
            "<p>问题描述</p>",
            True,
        ),
        BugDetailField(
            "DTS_RootCauseAnalysis",
            "根因分析",
            "<p>时序问题</p>",
            True,
        ),
        BugDetailField(
            "DTS_FixContent",
            "修复内容",
            "<p>增加状态同步</p>",
            True,
        ),
    )
    detail = BugDetail(
        workflow_key="1",
        object_oid="2",
        bug_number="Bug001",
        title="示例缺陷",
        description_html="<p>问题描述</p>",
        root_cause_html="<p>时序问题</p>",
        fix_content_html="<p>增加状态同步</p>",
        fields=fields,
    )
    markdown = build_retrospective_markdown(
        "2026-0730",
        "危国",
        [
            {
                "id": "Bug001",
                "severity": "2",
                "status": "关闭",
                "module": "公共模块 | PubSVC",
                "resolvedByName": "危国",
                "title": "示例缺陷",
            }
        ],
        {"Bug001": detail},
    )
    assert retrospective_filename("2026-0730", "危国") == "0730版本bug回溯-危国.md"
    assert "# 0730版本 bug回溯" in markdown
    assert "**周期**：0730版本SIT测试" in markdown
    assert "## 一、缺陷概况" in markdown
    assert "## 二、主要问题根因" in markdown
    assert "## 三、典型Bug深度分析" in markdown
    assert "**问题**：示例缺陷" in markdown
    assert "**根因**：时序问题" in markdown
    assert "**已修复**：增加状态同步" in markdown
    assert "## 四、后续改进" in markdown
    assert "时序问题" in markdown
    assert "增加状态同步" in markdown
    assert "**模块**：公共模块 | PubSVC" in markdown
    assert "Bug逐项回溯" not in markdown


def test_workload_import() -> None:
    from app import (
        BugStatisticsApp,
        DatePickerPopup,
        copy_workload_template,
        has_workload_access,
    )

    assert has_workload_access("T0423")
    assert has_workload_access("t0101")
    assert not has_workload_access("T9999")
    assert not has_workload_access("")

    with TemporaryDirectory() as template_directory:
        template_path = Path(template_directory) / "绩效工作量导入模板.xlsx"
        copy_workload_template(template_path)
        template_workbook = load_workbook(template_path, read_only=False, data_only=True)
        assert template_workbook.sheetnames == ["工作量导入", "填写说明"]
        assert list(template_workbook["工作量导入"].iter_rows(
            min_row=1,
            max_row=1,
            values_only=True,
        ))[0] == (
            "需求号",
            "工作描述",
            "计划开始时间",
            "计划结束时间",
            "工时/天",
            "责任人",
        )
        template_sheet = template_workbook["工作量导入"]
        assert template_sheet.max_row == 100
        assert template_sheet["A2"].border.left.style == "thin"
        assert template_sheet["A2"].border.bottom.style == "thin"
        assert template_sheet["F100"].border.right.style == "thin"
        assert template_sheet["F100"].border.bottom.style == "thin"
        assert template_sheet.row_dimensions[2].height == 23
        template_workbook.close()
        assert read_workload_excel(template_path, "2026-0830") == []

        app = BugStatisticsApp()
        app.withdraw()
        app.logged_in = True
        app.performance_staff_number = "T0423"
        downloaded_path = Path(template_directory) / "downloaded-template.xlsx"
        with (
            patch(
                "app.filedialog.asksaveasfilename",
                return_value=str(downloaded_path),
            ),
            patch("app.messagebox.showinfo"),
            patch.object(app, "_save_settings"),
        ):
            app._download_workload_template()
        assert app.workload_file_var.get() == str(downloaded_path)
        assert read_workload_excel(downloaded_path, "2026-0830") == []
        app.update()
        app.after_cancel(app._worker_poll_after_id)
        app.destroy()

    allocation = allocate_ui_workload(16, "普通前端任务")
    assert allocation["reqAnalyzeWorkLoad"] == 4
    assert allocation["keyControlsCount"] == 3
    assert allocation["selfTestingCount"] == 7
    assert calculate_workload_hours(allocation) == 16

    for hours in (0.1, 0.5, 1, 2, 4, 8, 12, 16, 19.2, 24, 32):
        generated = allocate_ui_workload(hours, "普通前端任务")
        assert calculate_workload_hours(generated) == hours
        assert generated["reqAnalyzeWorkLoad"] in {0, 2, 4, 6, 8}
        assert generated["preResearchWorkLoad"] >= 0
        assert generated["selfTestingCount"] >= 0

    fractional_ui = allocate_ui_workload(19.2, "普通前端任务")
    assert fractional_ui["preResearchWorkLoad"] == 0.2
    assert calculate_workload_hours(fractional_ui) == 19.2

    control_ui = allocate_ui_workload(12, "页面表格和树控件开发")
    assert control_ui["keyControlsCount"] == 3
    assert control_ui["useNewIntefaceCount"] == 0
    assert calculate_workload_hours(control_ui) == 12

    new_interface_ui = allocate_ui_workload(12, "新增接口开发和联调")
    assert new_interface_ui["useNewIntefaceCount"] == 3
    assert new_interface_ui["keyControlsCount"] == 1
    assert calculate_workload_hours(new_interface_ui) == 12

    mature_interface_ui = allocate_ui_workload(8, "成熟接口对接")
    assert mature_interface_ui["useMatureIntefaceCount"] > 0
    assert mature_interface_ui["useNewIntefaceCount"] == 0
    assert calculate_workload_hours(mature_interface_ui) == 8

    optional_ui = allocate_ui_workload(16, "局部刷新并封装通用组件")
    assert optional_ui["refreshUIControl"] is True
    assert optional_ui["needEncapsulatedComponent"] is True
    assert calculate_workload_hours(optional_ui) == 16

    immature_ui = allocate_ui_workload(12, "不成熟组件联调")
    assert immature_ui["useImmatureIntefaceCount"] == 1
    assert calculate_workload_hours(immature_ui) == 12

    research_ui = allocate_ui_workload(8, "技术预研和可行性验证")
    assert research_ui["preResearchWorkLoad"] >= 2
    assert calculate_workload_hours(research_ui) == 8

    assert allocation["selfTestingCount"] * 0.5 < 16 * 0.3

    server_allocation = allocate_server_workload(12)
    assert server_allocation["methodType"] == "全新开发"
    assert server_allocation["reqAnalyzeWorkLoad"] == 2
    assert server_allocation["codeComplexity"] == 4
    assert server_allocation["unitTestNumber"] == 6
    assert calculate_workload_hours(server_allocation) == 12
    for hours in (1, 2, 4, 8, 12, 16, 24, 32):
        generated = allocate_server_workload(hours, "普通后端任务")
        assert calculate_workload_hours(generated) == hours
        assert generated["reqAnalyzeWorkLoad"] in {0, 2, 4, 6, 8}
        assert generated["codeComplexity"] in {0, 1, 2, 4, 8, 16}

    encapsulation = allocate_server_workload(8, "接口封装")
    assert encapsulation["methodType"] == "业务封装"
    assert encapsulation["unitTestNumber"] == 4
    assert calculate_workload_hours(encapsulation) == 8
    database_task = allocate_server_workload(12, "数据库查询")
    assert database_task["isDBOperation"] is True
    assert calculate_workload_hours(database_task) == 12
    ui_debug_task = allocate_server_workload(12, "前端联调")
    assert ui_debug_task["isNeedUiDebug"] is True
    assert calculate_workload_hours(ui_debug_task) == 12
    existing_server_record = {
        "methodType": "全新开发",
        "isDBOperation": True,
        "dbAdapteNumer": 8,
        "codeComplexity": 4,
        "unitTestNumber": 8,
        "isNeedUiDebug": True,
    }
    assert calculate_workload_hours(existing_server_record) == 20.4

    workload_api = WorkloadApiClient(object())
    workload_api._post_json = lambda *args, **kwargs: {
        "data": {
            "children": [
                {
                    "workItemGroup": "开发一组",
                    "methods": [{"getUiDef": SERVER_WORKLOAD_DEFINITION}],
                },
                {
                    "workItemGroup": "开发一组-前端",
                    "methods": [{"getUiDef": UI_WORKLOAD_DEFINITION}],
                },
            ]
        }
    }
    groups = workload_api.get_groups()
    assert [group.name for group in groups] == ["开发一组", "开发一组-前端"]
    assert workload_api._definition_path("开发一组") == SERVER_WORKLOAD_DEFINITION

    with TemporaryDirectory() as directory:
        path = Path(directory) / "workload.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(
            [
                "需求号",
                "工作描述",
                "计划开始时间",
                "计划结束时间",
                "工时/天",
                "责任人",
            ]
        )
        sheet.append(
            [
                "SR001",
                "批量录入测试",
                "2026-08-10",
                "2026-08-11",
                2,
                "危国",
            ]
        )
        sheet.append(
            [
                "SR404",
                "不存在的需求",
                "2026-08-12",
                "2026-08-13",
                1,
                "危国",
            ]
        )
        workbook.save(path)
        workbook.close()

        items = read_workload_excel(path, "2026-0830")
        assert len(items) == 2
        assert items[0].plan_start_date == "2026-08-10"

        valid_content = path.read_bytes()
        with (
            patch(
                "workload_client.Path.read_bytes",
                side_effect=[b"incomplete save", valid_content],
            ),
            patch("workload_client.sleep"),
        ):
            retried_items = read_workload_excel(path, "2026-0830")
        assert len(retried_items) == 2

        with (
            patch(
                "workload_client.Path.read_bytes",
                side_effect=(
                    [b"partial save a", b"partial save b"]
                    + [b"partial save b"] * 29
                    + [valid_content]
                ),
            ),
            patch("workload_client.sleep") as changing_sleep,
        ):
            slowly_saved_items = read_workload_excel(path, "2026-0830")
        assert len(slowly_saved_items) == 2
        assert changing_sleep.call_count == 31

        invalid_path = Path(directory) / "invalid.xlsx"
        invalid_path.write_bytes(b"not an xlsx workbook")
        try:
            with (
                patch("workload_client.sleep") as mocked_sleep,
                patch(
                    "workload_client._read_open_office_xlsx_bytes",
                    return_value=None,
                ),
            ):
                read_workload_excel(invalid_path, "2026-0830")
        except Exception as exc:
            assert "不是有效的 .xlsx" in str(exc)
            assert "等待 15 秒" in str(exc)
            assert mocked_sleep.call_count == 30
        else:
            raise AssertionError("无效的 xlsx 文件应被拒绝")

        locked_path = Path(directory) / "locked.xlsx"
        locked_path.write_bytes(b"incomplete workbook")
        locked_path.with_name("~$locked.xlsx").write_bytes(b"office owner file")
        try:
            with (
                patch("workload_client.sleep") as locked_sleep,
                patch(
                    "workload_client._read_open_office_xlsx_bytes",
                    return_value=None,
                ),
            ):
                read_workload_excel(locked_path, "2026-0830")
        except Exception as exc:
            assert "仍在 WPS/Excel 中打开" in str(exc)
            assert "Ctrl+S" in str(exc)
            assert "不是有效的 .xlsx" not in str(exc)
            assert locked_sleep.call_count == 30
        else:
            raise AssertionError("打开中的 xlsx 应显示文件占用提示")

        intermittent_path = Path(directory) / "intermittent.xlsx"
        intermittent_path.write_bytes(b"incomplete workbook")
        try:
            with (
                patch(
                    "workload_client.Path.read_bytes",
                    side_effect=[PermissionError("locked")] + [b"incomplete"] * 60,
                ),
                patch("workload_client.sleep") as intermittent_sleep,
                patch(
                    "workload_client._read_open_office_xlsx_bytes",
                    return_value=None,
                ),
            ):
                read_workload_excel(intermittent_path, "2026-0830")
        except Exception as exc:
            assert "仍在 WPS/Excel 中打开" in str(exc)
            assert intermittent_sleep.call_count == 60
        else:
            raise AssertionError("间歇性文件锁不应被后续不完整读取掩盖")

        office_locked_path = Path(directory) / "office-locked.xlsx"
        office_locked_path.write_bytes(valid_content)
        with (
            patch(
                "workload_client.Path.read_bytes",
                side_effect=PermissionError("locked by WPS"),
            ),
            patch(
                "workload_client._read_open_office_xlsx_bytes",
                return_value=valid_content,
            ) as office_snapshot,
            patch("workload_client.sleep") as office_sleep,
        ):
            office_items = read_workload_excel(office_locked_path, "2026-0830")
        assert len(office_items) == 2
        assert office_snapshot.call_count == 1
        assert office_sleep.call_count == 0

        encrypted_path = Path(directory) / "encrypted.xlsx"
        encrypted_path.write_bytes(
            b"\xD0\xCF\x11\xE0\xA1\xB1\x1A\xE1" + b"EncryptedPackage"
        )
        try:
            with (
                patch("workload_client.sleep") as encrypted_sleep,
                patch(
                    "workload_client._read_open_office_xlsx_bytes",
                    return_value=None,
                ),
            ):
                read_workload_excel(encrypted_path, "2026-0830")
        except Exception as exc:
            assert "加密工作簿或旧版 .xls" in str(exc)
            assert "保持工作簿打开" in str(exc)
            assert encrypted_sleep.call_count == 0
        else:
            raise AssertionError("加密工作簿应显示独立提示")

        with patch(
            "workload_client._read_open_office_xlsx_bytes",
            return_value=valid_content,
        ) as decrypted_snapshot:
            encrypted_items = read_workload_excel(encrypted_path, "2026-0830")
        assert len(encrypted_items) == 2
        assert decrypted_snapshot.call_count == 1

        context = WorkloadContext(
            plan_version="2026-0830",
            module="开发一组-前端",
            load_path="/task/load",
            save_path="/task/save",
            load_params={"planVersion": "2026-0830", "module": "开发一组-前端"},
            requirements={"SR001": "测试需求"},
            developers={"危国": Developer("危国", "weig", "前端")},
            existing_tasks=[],
            next_ord=3,
        )
        preview = build_workload_preview(items, context, ALL_GROUP_MEMBERS)
        assert len(preview.submittable_rows) == 1
        assert preview.error_count == 1
        assert preview.submittable_rows[0].record["developerId"] == "weig"
        assert preview.computed_hours == 16

        app = BugStatisticsApp()
        app.withdraw()
        app.logged_in = True
        app.performance_staff_number = "T0423"
        app.workload_context = context
        app.workload_items = list(items)
        app.workload_preview = preview
        app.workload_selected_rows = {
            row.source.excel_row for row in preview.submittable_rows
        }
        app._render_workload_preview()
        app.update_idletasks()
        assert app.workload_tree.heading("requested_hours", "text") == "Excel 工时（小时）"
        assert app.workload_tree.heading("computed_hours", "text") == "计算工时（小时）"
        assert "工作量录入" in app.workload_nav_button.cget("text")
        assert app.workload_tree.set("2", "selected") == "☑"
        assert app.workload_tree.set("3", "selected") == "☐"
        assert len(app._selected_workload_rows()) == 1
        assert "已勾选 1 行" in app.workload_summary_var.get()
        checkbox, status_label, delete_button = app.workload_row_widgets["2"]
        assert checkbox.cget("cursor") == "hand2"
        assert "15" in str(checkbox.cget("font"))
        assert status_label.cget("foreground") == "#45E0B2"
        assert delete_button.cget("foreground") == "#FF7F96"
        assert delete_button.cget("cursor") == "hand2"
        error_checkbox, error_status, _error_delete = app.workload_row_widgets["3"]
        assert str(error_checkbox.cget("state")) == "disabled"
        assert error_checkbox.cget("cursor") == "no"
        assert error_status.cget("foreground") == "#FF7F96"
        assert app.workload_tree.heading("selected", "text") == ""
        assert app.workload_select_all_button.cget("text") == "☑"
        assert (
            str(app.workload_select_all_button.cget("font"))
            == str(checkbox.cget("font"))
        )
        app._toggle_all_workload_rows()
        assert not app.workload_selected_rows
        assert app.workload_select_all_button.cget("text") == "☐"
        app._toggle_all_workload_rows()
        assert app.workload_selected_rows == {2}
        app._update_workload_access()
        app.notebook.select(app.workload_tab)
        app._render_workload_preview()
        checkbox, _status_label, delete_button = app.workload_row_widgets["2"]
        app.deiconify()
        app.update()
        assert checkbox.winfo_ismapped()
        assert app.workload_select_all_button.winfo_ismapped()
        assert delete_button.winfo_ismapped()
        assert checkbox.place_info()
        assert delete_button.place_info()
        app.withdraw()

        selected_dates = []
        date_picker = DatePickerPopup(
            app,
            "2026-08-24",
            selected_dates.append,
            10,
            10,
        )
        assert date_picker.display_year == 2026
        assert date_picker.display_month == 8
        date_picker._select_day(25)
        app.update()
        assert selected_dates == ["2026-08-25"]

        app._add_workload_row()
        added_row = max(app.workload_added_rows)
        assert app.workload_add_dialog is None
        assert app.workload_tree.set(str(added_row), "excel_row") == "新增"
        assert app.workload_tree.set(str(added_row), "status") == "待填写"
        assert "还需填写" in app.workload_tree.set(str(added_row), "message")
        for pending_column in (
            "developer",
            "require_no",
            "task_name",
            "plan_start",
            "plan_finish",
            "requested_hours",
        ):
            assert (
                app.workload_tree.set(str(added_row), pending_column)
                == "待填写"
            )
        assert app.workload_tree.set(str(added_row), "computed_hours") == "—"
        assert added_row not in app.workload_selected_rows
        assert str(app.workload_tree.cget("selectmode")) == "none"
        assert str(app.workload_add_button.cget("style")) == "Outline.TButton"
        app.update_idletasks()
        app._close_workload_editor()
        app._open_workload_cell_editor(
            str(added_row),
            app._workload_column_id("developer"),
        )
        assert isinstance(app.workload_editor, ttk.Combobox)
        assert "危国" in app.workload_editor.cget("values")
        assert app.workload_editor.cget("style") == "WorkloadEditor.TCombobox"
        app._close_workload_editor()
        app._open_workload_cell_editor(
            str(added_row),
            app._workload_column_id("require_no"),
        )
        assert isinstance(app.workload_editor, ttk.Entry)
        assert not isinstance(app.workload_editor, ttk.Combobox)
        assert app.workload_editor.cget("style") == "WorkloadEditor.TEntry"
        assert app.workload_editor.get() == ""
        app.workload_editor.insert(0, "SR001")
        app._commit_workload_edit(
            added_row,
            "require_no",
            app.workload_editor.get(),
        )
        assert next(
            item
            for item in app.workload_items
            if item.excel_row == added_row
        ).require_no == "SR001"
        app._open_workload_cell_editor(
            str(added_row),
            app._workload_column_id("task_name"),
        )
        assert isinstance(app.workload_editor, ttk.Entry)
        assert not isinstance(app.workload_editor, ttk.Combobox)
        app._close_workload_editor()
        app._apply_workload_edit(added_row, "developer", "危国")
        app._apply_workload_edit(added_row, "task_name", "表格内新增测试")
        app._apply_workload_edit(added_row, "plan_start", "2026-08-24")
        app._apply_workload_edit(added_row, "plan_finish", "2026-08-25")
        app._apply_workload_edit(added_row, "requested_hours", "8")
        added_preview_row = next(
            row
            for row in app.workload_preview.rows
            if row.source.excel_row == added_row
        )
        assert added_preview_row.status == "可提交"
        assert added_preview_row.source.developer_name == "危国"
        assert added_preview_row.source.require_no == "SR001"
        assert added_preview_row.source.task_name == "表格内新增测试"
        assert added_preview_row.source.plan_start_date == "2026-08-24"
        assert added_preview_row.source.plan_finish_date == "2026-08-25"
        assert added_preview_row.source.requested_days == 1
        assert added_row in app.workload_selected_rows
        assert str(app.workload_add_button.cget("state")) == "normal"
        assert int(app.workload_template_button.grid_info()["column"]) == 9
        assert int(app.workload_add_button.grid_info()["column"]) == 10
        assert (
            app.workload_add_button.winfo_parent()
            == app.workload_template_button.winfo_parent()
        )
        app._delete_workload_row(added_row)
        assert added_row not in app.workload_added_rows
        assert all(item.excel_row != added_row for item in app.workload_items)

        app.workload_editor = ttk.Entry(app)
        app._commit_workload_edit(2, "requested_hours", "8")
        assert app.workload_items[0].requested_days == 1
        assert app.workload_tree.set("2", "requested_hours") == "8"
        assert app.workload_tree.set("2", "selected") == "☑"

        app.workload_selected_rows.clear()
        app._render_workload_preview()
        assert not app._selected_workload_rows()
        assert str(app.workload_submit_button["state"]) == "disabled"

        app.workload_selected_rows.add(2)
        submitted = {}

        def capture_submit(operation, function):
            submitted["operation"] = operation
            submitted["rows"] = function()

        with (
            patch("app.messagebox.askyesno", return_value=True),
            patch.object(app, "_run_worker", side_effect=capture_submit),
            patch.object(
                app.workload_api,
                "submit",
                side_effect=lambda _context, rows: list(rows),
            ),
        ):
            app._submit_workload()
        assert submitted["operation"] == "workload_submit"
        assert len(submitted["rows"]) == 1
        assert submitted["rows"][0].source.excel_row == 2

        app._delete_workload_row(2)
        assert all(item.excel_row != 2 for item in app.workload_items)
        assert "2" not in app.workload_tree.get_children()
        app.update()
        app.after_cancel(app._worker_poll_after_id)
        app.destroy()

        server_context = WorkloadContext(
            plan_version="2026-0830",
            module="开发一组",
            load_path="/task/getSvrDevTaskByPlanAndModule",
            save_path="/task/batchSaveServerDevTasks",
            load_params={"planVersion": "2026-0830", "module": "开发一组"},
            requirements={"SR001": "测试需求"},
            developers={"危国": Developer("危国", "weig", "后端")},
            existing_tasks=[],
            next_ord=1,
            ui_definition_path=SERVER_WORKLOAD_DEFINITION,
        )
        server_preview = build_workload_preview(
            [items[0]], server_context, ALL_GROUP_MEMBERS
        )
        assert len(server_preview.submittable_rows) == 1
        server_record = server_preview.submittable_rows[0].record
        assert server_record is not None
        assert server_record["methodType"] == "全新开发"
        assert server_record["reqAnalyzeWorkLoad"] == 2
        assert server_record["codeComplexity"] == 8
        assert server_record["unitTestNumber"] == 6
        assert server_preview.computed_hours == 16

        duplicate_record = dict(preview.submittable_rows[0].record)
        duplicate_context = WorkloadContext(
            plan_version=context.plan_version,
            module=context.module,
            load_path=context.load_path,
            save_path=context.save_path,
            load_params=context.load_params,
            requirements=context.requirements,
            developers=context.developers,
            existing_tasks=[duplicate_record],
            next_ord=4,
        )
        duplicate_preview = build_workload_preview(
            [items[0]], duplicate_context, ALL_GROUP_MEMBERS
        )
        assert duplicate_preview.duplicate_count == 1
        assert not duplicate_preview.submittable_rows
        assert duplicate_preview.rows[0].record is None
        assert "不会重复提交" in duplicate_preview.rows[0].message


if __name__ == "__main__":
    test_analytics()
    test_crm_helpers()
    test_credential_store()
    test_all_detail_fields()
    test_crm_search_helpers()
    test_retrospective_markdown()
    test_workload_import()
    print("核心逻辑测试通过")
