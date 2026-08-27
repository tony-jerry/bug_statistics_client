"""绩效工作量 Excel 解析、动态接口发现与批量保存。"""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date, datetime
from io import BytesIO
from math import ceil
from pathlib import Path
from tempfile import TemporaryDirectory
from time import sleep, time
from typing import Any, Iterable
from uuid import uuid4
from zipfile import BadZipFile, ZipFile
from zlib import crc32
from xml.sax.saxutils import escape

import requests
from openpyxl import load_workbook

try:
    import pythoncom
    from win32com.client import GetActiveObject
except ImportError:  # pragma: no cover - only relevant when running from source
    pythoncom = None
    GetActiveObject = None

from api_client import AuthenticationError, BugApiClient, BugClientError


ALL_GROUP_MEMBERS = "全部组员"
DEFAULT_WORKLOAD_GROUP = "开发一组-前端"
UI_WORKLOAD_DEFINITION = "/uiDef/getUIWorkLoadTableDef"
SERVER_WORKLOAD_DEFINITION = "/uiDef/getServerWorkLoadTableDef"
SUPPORTED_WORKLOAD_DEFINITIONS = {
    UI_WORKLOAD_DEFINITION,
    SERVER_WORKLOAD_DEFINITION,
}
_OFFICE_SNAPSHOT_DIAGNOSTICS: list[str] = []


@dataclass(frozen=True)
class WorkloadGroup:
    name: str
    ui_definition_path: str


@dataclass(frozen=True)
class Developer:
    name: str
    user_id: str
    job: str = ""


@dataclass(frozen=True)
class ExcelWorkItem:
    excel_row: int
    require_no: str
    task_name: str
    plan_start_date: str
    plan_finish_date: str
    requested_days: float
    developer_name: str


@dataclass
class WorkloadPreviewRow:
    source: ExcelWorkItem
    status: str
    message: str
    computed_hours: float
    record: dict[str, Any] | None = None

    @property
    def is_submittable(self) -> bool:
        return self.record is not None and self.status in {"可提交", "提醒"}


@dataclass
class WorkloadPreview:
    rows: list[WorkloadPreviewRow]
    source_row_count: int
    skipped_row_count: int
    ignored_developers: tuple[str, ...] = ()

    @property
    def submittable_rows(self) -> list[WorkloadPreviewRow]:
        return [row for row in self.rows if row.is_submittable]

    @property
    def error_count(self) -> int:
        return sum(row.status == "错误" for row in self.rows)

    @property
    def duplicate_count(self) -> int:
        return sum(row.status == "已存在" for row in self.rows)

    @property
    def warning_count(self) -> int:
        return sum(row.status == "提醒" for row in self.rows)

    @property
    def requested_hours(self) -> float:
        return round(
            sum(row.source.requested_days * 8 for row in self.submittable_rows), 2
        )

    @property
    def computed_hours(self) -> float:
        return round(sum(row.computed_hours for row in self.submittable_rows), 2)


@dataclass
class WorkloadContext:
    plan_version: str
    module: str
    load_path: str
    save_path: str
    load_params: dict[str, Any]
    requirements: dict[str, str]
    developers: dict[str, Developer]
    existing_tasks: list[dict[str, Any]]
    next_ord: int
    ui_definition_path: str = UI_WORKLOAD_DEFINITION


@dataclass(frozen=True)
class WorkloadSubmitResult:
    submitted_count: int
    verified_count: int
    before_count: int
    after_count: int
    message: str


HEADER_ALIASES = {
    "require_no": ("需求号", "需求编号"),
    "task_name": ("工作描述", "任务描述"),
    "plan_start_date": ("计划开始时间", "计划开始", "开始时间", "预计开始"),
    "plan_finish_date": (
        "计划结束时间",
        "计划完成时间",
        "计划结束",
        "计划完成",
        "结束时间",
        "预计结束",
    ),
    "requested_days": ("工时/天", "工作量/天", "人天", "计划人天"),
    "developer_name": ("责任人", "开发资源", "开发人员"),
}


def _text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value)).strip()
    return str(value).strip()


def _parse_date(value: Any, default_year: int) -> str:
    if value is None or _text(value) == "":
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()

    raw = _text(value)
    normalized = raw.replace("/", "-")
    for pattern in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(normalized, pattern).date().isoformat()
        except ValueError:
            pass
    if "." in raw:
        parts = raw.split(".")
        if len(parts) == 2 and all(part.strip().isdigit() for part in parts):
            try:
                return date(default_year, int(parts[0]), int(parts[1])).isoformat()
            except ValueError:
                return raw
    return raw


def _find_header_indexes(header: Iterable[Any]) -> dict[str, int]:
    normalized = {_text(value): index for index, value in enumerate(header) if _text(value)}
    indexes: dict[str, int] = {}
    missing: list[str] = []
    for field_name, aliases in HEADER_ALIASES.items():
        matched = next((normalized[name] for name in aliases if name in normalized), None)
        if matched is None:
            missing.append(aliases[0])
        else:
            indexes[field_name] = matched
    if missing:
        raise BugClientError("Excel 缺少必需列：" + "、".join(missing))
    return indexes


def _valid_xlsx_snapshot(content: bytes) -> bool:
    if not content.startswith(b"PK"):
        return False
    try:
        with ZipFile(BytesIO(content)) as archive:
            names = set(archive.namelist())
            if "[Content_Types].xml" not in names or "xl/workbook.xml" not in names:
                return False
            return archive.testzip() is None
    except (BadZipFile, OSError):
        return False


def _office_lock_path(excel_path: Path) -> Path:
    """Return the owner-file path created while Excel/WPS keeps a workbook open."""

    return excel_path.with_name(f"~${excel_path.name}")


def office_snapshot_diagnostics() -> tuple[str, ...]:
    return tuple(_OFFICE_SNAPSHOT_DIAGNOSTICS)


def _current_office_lock_exists(excel_path: Path) -> bool:
    """Ignore abandoned Office owner files left behind by old WPS sessions."""

    try:
        lock_age_seconds = time() - _office_lock_path(excel_path).stat().st_mtime
        return -60 <= lock_age_seconds <= 24 * 60 * 60
    except OSError:
        return False


def _xlsx_bytes_from_open_workbook_values(workbook: Any) -> bytes | None:
    """Build a plain in-memory xlsx from the first open sheet's displayed values."""

    try:
        source_sheet = workbook.Worksheets.Item(1)
        used_range = source_sheet.UsedRange
        last_row = int(used_range.Row) + int(used_range.Rows.Count) - 1
        last_column = int(used_range.Column) + int(used_range.Columns.Count) - 1
        if last_row < 1 or last_column < 1:
            return None
        # The import template has six columns.  Keep a generous cap to avoid a
        # corrupted UsedRange allocating an unbounded COM matrix.
        last_column = min(last_column, 256)
        values = source_sheet.Range(
            source_sheet.Cells(1, 1),
            source_sheet.Cells(last_row, last_column),
        ).Value

        if last_row == 1 and last_column == 1:
            rows = ((values,),)
        elif isinstance(values, tuple):
            rows = values
        else:
            return None

        def column_name(column: int) -> str:
            name = ""
            while column:
                column, remainder = divmod(column - 1, 26)
                name = chr(65 + remainder) + name
            return name

        def xml_text(value: Any) -> str:
            raw = str(value)
            legal = "".join(
                character
                for character in raw
                if character in "\t\n\r"
                or "\x20" <= character <= "\ud7ff"
                or "\ue000" <= character <= "\ufffd"
            )
            return escape(legal)

        row_xml: list[str] = []
        for row_number, row in enumerate(rows, start=1):
            row_values = row if isinstance(row, tuple) else (row,)
            cell_xml: list[str] = []
            for column_number, value in enumerate(row_values, start=1):
                if value is None:
                    continue
                reference = f"{column_name(column_number)}{row_number}"
                if isinstance(value, bool):
                    cell_xml.append(
                        f'<c r="{reference}" t="b"><v>{int(value)}</v></c>'
                    )
                elif isinstance(value, (int, float)):
                    cell_xml.append(f'<c r="{reference}" t="n"><v>{value}</v></c>')
                else:
                    cell_xml.append(
                        f'<c r="{reference}" t="inlineStr"><is><t xml:space="preserve">'
                        f"{xml_text(value)}</t></is></c>"
                    )
            if cell_xml:
                row_xml.append(f'<row r="{row_number}">{"".join(cell_xml)}</row>')

        last_cell = f"{column_name(last_column)}{last_row}"
        sheet_xml = (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
            f'<dimension ref="A1:{last_cell}"/><sheetData>{"".join(row_xml)}</sheetData>'
            "</worksheet>"
        )
        stream = BytesIO()
        with ZipFile(stream, "w") as archive:
            archive.writestr(
                "[Content_Types].xml",
                '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
                '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
                '<Default Extension="xml" ContentType="application/xml"/>'
                '<Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>'
                '<Override PartName="/xl/worksheets/sheet1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>'
                "</Types>",
            )
            archive.writestr(
                "_rels/.rels",
                '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
                '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/>'
                "</Relationships>",
            )
            archive.writestr(
                "xl/workbook.xml",
                '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                '<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
                'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
                '<sheets><sheet name="Sheet1" sheetId="1" r:id="rId1"/></sheets>'
                "</workbook>",
            )
            archive.writestr(
                "xl/_rels/workbook.xml.rels",
                '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
                '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet1.xml"/>'
                "</Relationships>",
            )
            archive.writestr("xl/worksheets/sheet1.xml", sheet_xml)
        content = stream.getvalue()
        return content if _valid_xlsx_snapshot(content) else None
    except Exception as exc:
        _OFFICE_SNAPSHOT_DIAGNOSTICS.append(f"value snapshot failed: {exc!r}")
        return None


def _read_open_office_xlsx_bytes(excel_path: Path) -> bytes | None:
    """Ask an already-running Excel/WPS instance for a read-only disk snapshot."""

    if pythoncom is None or GetActiveObject is None:
        return None

    _OFFICE_SNAPSHOT_DIAGNOSTICS.clear()
    pythoncom.CoInitialize()
    try:
        target_path = excel_path.resolve(strict=False)
        _OFFICE_SNAPSHOT_DIAGNOSTICS.append(f"target={target_path}")
        for prog_id in ("Ket.Application", "Excel.Application"):
            try:
                application = GetActiveObject(prog_id)
                workbook_count = int(application.Workbooks.Count)
                _OFFICE_SNAPSHOT_DIAGNOSTICS.append(
                    f"{prog_id}: workbook_count={workbook_count}"
                )
            except Exception as exc:
                _OFFICE_SNAPSHOT_DIAGNOSTICS.append(
                    f"{prog_id}: unavailable: {exc!r}"
                )
                continue

            for index in range(1, workbook_count + 1):
                workbook = None
                try:
                    workbook = application.Workbooks.Item(index)
                    workbook_path = Path(str(workbook.FullName)).resolve(strict=False)
                    _OFFICE_SNAPSHOT_DIAGNOSTICS.append(
                        f"{prog_id}[{index}]={workbook_path}"
                    )
                    if workbook_path != target_path:
                        continue
                    with TemporaryDirectory(prefix="workload_excel_") as directory:
                        snapshot_path = Path(directory) / excel_path.name
                        # SaveCopyAs does not change the open workbook or its current path.
                        workbook.SaveCopyAs(str(snapshot_path))
                        content = snapshot_path.read_bytes()
                    if _valid_xlsx_snapshot(content):
                        return content
                    # Password/enterprise encryption may be preserved by
                    # SaveCopyAs.  Values are already decrypted in the open WPS
                    # workbook, so rebuild a plain temporary workbook from them.
                    return _xlsx_bytes_from_open_workbook_values(workbook)
                except Exception as exc:
                    _OFFICE_SNAPSHOT_DIAGNOSTICS.append(
                        f"{prog_id}[{index}] SaveCopyAs failed: {exc!r}"
                    )
                    value_snapshot = (
                        _xlsx_bytes_from_open_workbook_values(workbook)
                        if workbook is not None
                        else None
                    )
                    if value_snapshot is not None:
                        return value_snapshot
                    continue
        return None
    finally:
        pythoncom.CoUninitialize()


def _read_stable_xlsx_bytes(
    excel_path: Path,
    *,
    attempts: int = 31,
    delay_seconds: float = 0.5,
) -> bytes:
    """避开 WPS/Excel 后台保存期间短暂出现的不完整文件快照。"""

    base_attempts = max(attempts, 1)
    attempt_limit = base_attempts
    # When WPS is demonstrably replacing the file underneath us, allow another
    # 15 seconds.  A static invalid file still fails after the original window.
    changing_file_attempt_limit = base_attempts + max(attempts - 1, 0)
    attempt = 0
    sleep_count = 0
    previous_snapshot: tuple[int, int] | None = None
    last_error: OSError | None = None
    saw_permission_error = False
    office_snapshot_attempted = False
    last_size = 0
    while attempt < attempt_limit:
        try:
            content = excel_path.read_bytes()
            last_size = len(content)
            last_error = None
            if _valid_xlsx_snapshot(content):
                return content
            if content.startswith(b"\xD0\xCF\x11\xE0\xA1\xB1\x1A\xE1"):
                office_content = _read_open_office_xlsx_bytes(excel_path)
                if office_content is not None:
                    return office_content
                raise BugClientError(
                    "Excel 文件是加密工作簿或旧版 .xls 格式；请先在 "
                    "WPS/Excel 中输入密码并保持工作簿打开，然后重新加载。"
                    "也可取消打开密码后另存为标准 .xlsx"
                )
            # WPS can temporarily replace an unlocked encrypted workbook with a
            # proprietary non-ZIP image that has neither the PK nor OLE magic.
            # Match the workbook by its open path rather than its disk signature.
            if not office_snapshot_attempted:
                office_snapshot_attempted = True
                office_content = _read_open_office_xlsx_bytes(excel_path)
                if office_content is not None:
                    return office_content
            current_snapshot = (last_size, crc32(content))
            snapshot_changed = (
                previous_snapshot is not None
                and current_snapshot != previous_snapshot
            )
            previous_snapshot = current_snapshot
            if snapshot_changed or saw_permission_error:
                attempt_limit = max(attempt_limit, changing_file_attempt_limit)
            if _current_office_lock_exists(excel_path) and not office_snapshot_attempted:
                office_snapshot_attempted = True
                office_content = _read_open_office_xlsx_bytes(excel_path)
                if office_content is not None:
                    return office_content
        except OSError as exc:
            last_error = exc
            saw_permission_error = saw_permission_error or isinstance(
                exc, PermissionError
            )
            if isinstance(exc, PermissionError) and not office_snapshot_attempted:
                office_snapshot_attempted = True
                office_content = _read_open_office_xlsx_bytes(excel_path)
                if office_content is not None:
                    return office_content
        attempt += 1
        if attempt < attempt_limit:
            sleep(delay_seconds)
            sleep_count += 1

    if office_snapshot_attempted:
        office_content = _read_open_office_xlsx_bytes(excel_path)
        if office_content is not None:
            return office_content

    # WPS may alternate between an exclusive lock and a readable but incomplete
    # snapshot while saving.  Do not let a later incomplete read hide the lock.
    if (
        isinstance(last_error, PermissionError)
        or saw_permission_error
        or _current_office_lock_exists(excel_path)
    ):
        raise BugClientError(
            f"无法读取 {excel_path.name}：检测到该文件仍在 WPS/Excel 中打开"
            "或正在保存。请先按 Ctrl+S 保存，再关闭该工作簿后重新加载"
        ) from last_error
    if last_error is not None:
        raise BugClientError(f"无法读取 Excel 文件：{last_error}") from last_error
    waited_seconds = sleep_count * max(delay_seconds, 0)
    raise BugClientError(
        f"等待 {waited_seconds:g} 秒后仍无法读取 {excel_path.name}"
        f"（当前 {last_size} 字节）。文件保存尚未完成，或不是有效的 .xlsx 工作簿；"
        "请确认选择的是刚保存的文件，仍失败时请在 WPS/Excel 中另存为 .xlsx"
    )


def read_workload_excel(path: str | Path, plan_version: str) -> list[ExcelWorkItem]:
    """读取工作量 Excel；空行会跳过，字段错误留给预览阶段展示。"""

    excel_path = Path(path)
    if not excel_path.is_file():
        raise BugClientError(f"Excel 文件不存在：{excel_path}")
    if excel_path.suffix.lower() != ".xlsx":
        raise BugClientError("仅支持 .xlsx 文件")
    try:
        default_year = int(plan_version[:4])
    except (TypeError, ValueError):
        default_year = datetime.now().year

    excel_bytes = _read_stable_xlsx_bytes(excel_path)

    try:
        # 使用内存快照解析，避免打包环境下文件路径句柄与 WPS 保存过程冲突。
        workbook = load_workbook(
            BytesIO(excel_bytes),
            read_only=True,
            data_only=True,
        )
    except BadZipFile as exc:
        raise BugClientError(
            "Excel 文件不是有效的 .xlsx 工作簿，请在 WPS/Excel 中另存为 .xlsx 后重试"
        ) from exc
    except Exception as exc:
        raise BugClientError(f"无法解析 Excel：{exc}") from exc

    try:
        sheet = workbook.active
        values = sheet.iter_rows(values_only=True)
        try:
            header = next(values)
        except StopIteration as exc:
            raise BugClientError("Excel 中没有数据") from exc
        indexes = _find_header_indexes(header)
        items: list[ExcelWorkItem] = []
        for excel_row, values_row in enumerate(values, start=2):
            values_list = list(values_row)

            def value(field_name: str) -> Any:
                index = indexes[field_name]
                return values_list[index] if index < len(values_list) else None

            raw_values = [value(field_name) for field_name in indexes]
            if all(_text(item) == "" for item in raw_values):
                continue
            days_raw = value("requested_days")
            try:
                requested_days = float(days_raw) if _text(days_raw) else 0.0
            except (TypeError, ValueError):
                requested_days = -1.0
            items.append(
                ExcelWorkItem(
                    excel_row=excel_row,
                    require_no=_text(value("require_no")),
                    task_name=_text(value("task_name")),
                    plan_start_date=_parse_date(value("plan_start_date"), default_year),
                    plan_finish_date=_parse_date(value("plan_finish_date"), default_year),
                    requested_days=requested_days,
                    developer_name=_text(value("developer_name")),
                )
            )
        return items
    finally:
        workbook.close()


def calculate_workload_hours(record: dict[str, Any]) -> float:
    if "methodType" in record:
        return calculate_server_workload_hours(record)
    return round(
        float(record.get("preResearchWorkLoad") or 0)
        + float(record.get("reqAnalyzeWorkLoad") or 0)
        + float(record.get("keyControlsCount") or 0) * 2
        + float(record.get("useMatureIntefaceCount") or 0) * 0.5
        + float(record.get("useNewIntefaceCount") or 0) * 2
        + float(record.get("useImmatureIntefaceCount") or 0) * 4
        + (4 if record.get("refreshUIControl") else 0)
        + (4 if record.get("needEncapsulatedComponent") else 0)
        + float(record.get("selfTestingCount") or 0) * 0.5,
        2,
    )


def calculate_server_workload_hours(record: dict[str, Any]) -> float:
    """与绩效系统前端的 serverWorkload 公式保持一致。"""

    working_hours = 0.0
    if record.get("methodType") == "全新开发":
        workload_factor = 0.0
        if record.get("isDBOperation"):
            workload_factor = 1.0
            table_count = float(record.get("tableCount") or 0)
            workload_factor *= 1 + (0.5 if table_count >= 5 else table_count * 0.1)
            if record.get("isIteratedModel"):
                workload_factor *= 1.5
            if record.get("isLinkLoop"):
                workload_factor *= 1.3
            if record.get("isIncludeExtProp"):
                workload_factor *= 1.3
            if record.get("needHighPerformance"):
                workload_factor *= 1.3
            db_count = float(record.get("dbAdapteNumer") or 0)
            if db_count > 0:
                workload_factor *= 1 + (0.6 if db_count >= 3 else db_count * 0.2)
        working_hours = 4 * workload_factor
        if record.get("isNeedUiDebug"):
            working_hours += 2
        working_hours += float(record.get("unitTestNumber") or 0)
    else:
        if record.get("isNeedUiDebug"):
            working_hours += 1
        working_hours += float(record.get("unitTestNumber") or 0) * 0.5
    working_hours += 4 if record.get("isImmatureComponentDebug") else 0
    working_hours += 4 if record.get("isNeedPrepareData") else 0
    working_hours += 0.5 if record.get("isDesignUML") else 0
    working_hours += float(record.get("reqAnalyzeWorkLoad") or 0)
    working_hours += float(record.get("codeComplexity") or 0)
    return ceil((working_hours - 1e-12) * 100) / 100


def _stable_percentage_bucket(allocation_key: str) -> int:
    """Map a stable task key into a repeatable percentage bucket."""
    return crc32(_text(allocation_key).casefold().encode("utf-8")) % 100


def allocate_ui_workload(
    total_hours: float,
    task_name: str = "",
    allocation_key: str = "",
) -> dict[str, Any]:
    """按前端历史比例和工作描述生成总量精确的字段组合。"""

    total_hours = max(round(float(total_hours), 2), 0.0)
    values: dict[str, Any] = {
        "preResearchWorkLoad": 0.0,
        "reqAnalyzeWorkLoad": 0.0,
        "keyControlsCount": 0.0,
        "useMatureIntefaceCount": 0.0,
        "useNewIntefaceCount": 0.0,
        "useImmatureIntefaceCount": 0.0,
        "refreshUIControl": False,
        "needEncapsulatedComponent": False,
        "selfTestingCount": 0.0,
    }
    if total_hours <= 0:
        return values

    normalized_name = _text(task_name).casefold()

    def contains_any(keywords: tuple[str, ...]) -> bool:
        return any(keyword.casefold() in normalized_name for keyword in keywords)

    is_pre_research = contains_any(
        (
            "预研",
            "技术调研",
            "可行性",
            "技术验证",
            "技术方案",
            "方案设计",
            "技术选型",
            "原型验证",
        )
    )
    needs_refresh = contains_any(
        ("不大刷", "局部刷新", "增量刷新", "刷新优化", "免刷新")
    )
    needs_component = contains_any(
        ("通用组件", "公共组件", "组件封装", "封装组件", "抽取组件", "组件化")
    )
    uses_immature_component = contains_any(
        ("不成熟组件", "新组件联调", "第三方组件联调", "三方组件联调", "组件调试")
    )
    mentions_control = contains_any(
        (
            "页面",
            "界面",
            "控件",
            "表格",
            "左侧树",
            "树展示",
            "菜单",
            "按钮",
            "弹窗",
            "表单",
            "列表",
            "布局",
            "交互",
            "单元格",
            "勾选框",
            "动画",
            "展示",
        )
    )
    mentions_interface = contains_any(("接口", "api", "联调", "对接"))
    mentions_immature_interface = contains_any(("不成熟接口", "新组件接口"))
    mentions_mature_interface = not mentions_immature_interface and contains_any(
        ("成熟接口", "已有接口", "复用接口")
    )
    mentions_new_interface = contains_any(
        ("新接口", "新增接口", "接口开发", "接口改造")
    )

    # 只有描述明确命中时才启用 4h 附加项，并始终预留至少 0.5h 自测。
    remaining = total_hours
    minimum_test_hours = 0.5 if total_hours >= 0.5 else 0.0
    for field_name, enabled in (
        ("refreshUIControl", needs_refresh),
        ("needEncapsulatedComponent", needs_component),
    ):
        if enabled and remaining >= 4.0 + minimum_test_hours:
            values[field_name] = True
            remaining = round(remaining - 4.0, 2)
    if uses_immature_component and remaining >= 4.0 + minimum_test_hours:
        values["useImmatureIntefaceCount"] = 1.0
        remaining = round(remaining - 4.0, 2)

    # 2026-01~07 八个前端组约 1983 条记录的可信核心贡献比例约为
    # 需求 16% / 主要控件 26% / 成熟接口 3% / 新接口 10% / 自测 17%。
    weights = {
        "requirements": 0.16,
        "controls": 0.26,
        "mature": 0.03,
        "new": 0.10,
        "testing": 0.17,
    }
    if mentions_control and not mentions_interface:
        weights.update(controls=0.42, mature=0.02, new=0.02, testing=0.18)
    elif mentions_interface and not mentions_control:
        weights.update(controls=0.08, mature=0.10, new=0.28, testing=0.18)
    if mentions_new_interface:
        weights.update(controls=0.08, mature=0.02, new=0.38, testing=0.18)
    elif mentions_mature_interface:
        weights.update(controls=0.08, mature=0.30, new=0.08, testing=0.18)

    ui_key = allocation_key or f"{task_name}|{total_hours:.2f}"
    meets_research_threshold = total_hours >= 16.0
    probabilistic_research = (
        meets_research_threshold
        and _stable_percentage_bucket(f"{ui_key}|pre-research") < 50
    )
    # 2 天是预研的硬门槛：低于 16h 时，即使描述命中预研关键词也不分配。
    # 达到门槛的普通任务稳定地取约 50%；明确的预研任务必定分配。
    # 预研字段最多 4h，并预留至少 0.5h 自测。
    if (
        meets_research_threshold
        and (is_pre_research or probabilistic_research)
        and remaining > minimum_test_hours
    ):
        target_research_hours = (
            min(round(total_hours * 0.25 * 2) / 2, 4.0)
            if is_pre_research
            else 4.0
        )
        research_hours = min(
            target_research_hours,
            remaining - minimum_test_hours,
            4.0,
        )
        values["preResearchWorkLoad"] = max(round(research_hours, 2), 0.0)
        remaining = round(remaining - values["preResearchWorkLoad"], 2)

    half_hour_units = max(int((remaining + 1e-9) / 0.5), 0)
    tail_hours = round(remaining - half_hour_units * 0.5, 2)
    # 小于 0.5h 的尾数属于自测，不能为了凑总工时写进预研字段。
    testing_tail_hours = tail_hours
    core_hours = half_hour_units * 0.5
    if half_hour_units == 0:
        values["selfTestingCount"] = round(testing_tail_hours / 0.5, 2)
        return values

    weight_total = sum(weights.values())
    targets = {
        field_name: core_hours * weight / weight_total
        for field_name, weight in weights.items()
    }
    requirement_values = (0.0, 2.0, 4.0, 6.0, 8.0)
    scale = max(core_hours, 1.0)
    best: tuple[tuple[float, ...], dict[str, float]] | None = None

    for requirement_hours in requirement_values:
        requirement_units = int(requirement_hours * 2)
        minimum_test_units = 1 if half_hour_units >= 1 else 0
        if requirement_units + minimum_test_units > half_hour_units:
            continue
        units_after_requirement = half_hour_units - requirement_units
        for control_count in range(units_after_requirement // 4 + 1):
            units_after_controls = units_after_requirement - control_count * 4
            for new_count in range(units_after_controls // 4 + 1):
                flexible_units = units_after_controls - new_count * 4
                target_testing_units = targets["testing"] / 0.5
                target_mature_units = targets["mature"] / 0.5
                ideal_testing_units = round(
                    (target_testing_units + flexible_units - target_mature_units) / 2
                )
                testing_units = min(
                    max(ideal_testing_units, minimum_test_units),
                    flexible_units,
                )
                mature_units = flexible_units - testing_units
                contributions = {
                    "requirements": requirement_hours,
                    "controls": control_count * 2.0,
                    "mature": mature_units * 0.5,
                    "new": new_count * 2.0,
                    "testing": testing_units * 0.5,
                }
                score = sum(
                    ((contributions[field] - targets[field]) / scale) ** 2
                    for field in targets
                )
                if mentions_control and core_hours >= 2 and control_count == 0:
                    score += 0.2
                if mentions_new_interface and core_hours >= 2 and new_count == 0:
                    score += 0.2
                if mentions_mature_interface and core_hours >= 0.5 and mature_units == 0:
                    score += 0.2
                candidate = {
                    "reqAnalyzeWorkLoad": requirement_hours,
                    "keyControlsCount": float(control_count),
                    "useMatureIntefaceCount": float(mature_units),
                    "useNewIntefaceCount": float(new_count),
                    "selfTestingCount": float(testing_units),
                }
                rank = (
                    score,
                    abs(requirement_hours - targets["requirements"]),
                    abs(contributions["testing"] - targets["testing"]),
                    float(control_count + new_count + mature_units + testing_units),
                )
                if best is None or rank < best[0]:
                    best = (rank, candidate)

    if best is not None:
        values.update(best[1])
    if testing_tail_hours:
        values["selfTestingCount"] = round(
            values["selfTestingCount"] + testing_tail_hours / 0.5,
            2,
        )

    calculated_hours = calculate_workload_hours(values)
    rounding_delta = round(total_hours - calculated_hours, 2)
    if abs(rounding_delta) > 0.001:
        values["selfTestingCount"] = round(
            values["selfTestingCount"] + rounding_delta / 0.5,
            2,
        )
    return values


def allocate_server_workload(
    total_hours: float,
    task_name: str = "",
    allocation_key: str = "",
) -> dict[str, Any]:
    """按后端高频字段生成可解释、优先与目标总量精确匹配的字段组合。"""

    total_hours = max(round(float(total_hours), 2), 0.0)
    normalized_name = _text(task_name).casefold()

    def contains_any(keywords: tuple[str, ...]) -> bool:
        return any(keyword.casefold() in normalized_name for keyword in keywords)

    method_type = (
        "业务封装"
        if contains_any(("业务封装", "接口封装", "透传", "转发", "代理"))
        else "全新开发"
    )
    values: dict[str, Any] = {
        "methodType": method_type,
        "reqAnalyzeWorkLoad": 0.0,
        "isDesignUML": False,
        "unitTestNumber": 0.0,
        "isNeedPrepareData": False,
        "isNeedUiDebug": False,
        "codeComplexity": 0.0,
        "isImmatureComponentDebug": False,
        "isDBOperation": False,
        "tableCount": 0.0,
        "isIteratedModel": False,
        "isLinkLoop": False,
        "isIncludeExtProp": False,
        "dbAdapteNumer": 0.0,
        "needHighPerformance": False,
    }

    # 后端录入只主动填写高频字段，需求分析保持为 0。
    uses_database = contains_any(
        ("数据库", "数据表", "表结构", "建表", "表脚本", "sql", "持久化", "dao")
    )
    adapts_database = contains_any(
        (
            "适配数据库",
            "数据库适配",
            "兼容数据库",
            "数据库兼容",
            "多数据库",
            "多库适配",
            "国产数据库",
        )
    )
    database_key = allocation_key or f"{task_name}|{total_hours:.2f}"
    # 明确的数据库任务必选；其余全新开发任务稳定地取 40% 分桶。
    # 以任务内容作为种子，因此重复预览不会随机改变已分配字段。
    probabilistic_database = (
        method_type == "全新开发"
        and total_hours >= 4.0
        and _stable_percentage_bucket(database_key) < 40
    )
    if method_type == "全新开发" and (uses_database or probabilistic_database):
        values["isDBOperation"] = True
    # 在已勾选数据库的记录中稳定地取 80% 填写适配数量；描述明确包含
    # 数据库适配/兼容的任务则必填。历史记录中 4 最常见，默认填 4。
    fills_database_adapter = (
        values["isDBOperation"]
        and _stable_percentage_bucket(f"{database_key}|adapter") < 80
    )
    if values["isDBOperation"] and (adapts_database or fills_database_adapter):
        values["dbAdapteNumer"] = 4.0
    ui_debug_hours = 2.0 if method_type == "全新开发" else 1.0
    semantic_hours = calculate_server_workload_hours(values)
    # 前端联调是后端记录的高频项，默认勾选；仅当总工时连该项本身都
    # 无法容纳时保持未勾选，避免生成的计算工时超过 Excel 工时。
    if semantic_hours + ui_debug_hours <= total_hours + 1e-9:
        values["isNeedUiDebug"] = True
        semantic_hours = calculate_server_workload_hours(values)

    remaining = max(round(total_hours - semantic_hours, 2), 0.0)
    complexity_values = (0.0, 1.0, 2.0, 4.0, 8.0, 16.0)
    # 原历史比例为需求分析 19% / 代码复杂度 45% / 单元测试 36%。
    # 需求分析不再自动填写后，将其余两项归一化为约 56% / 44%。
    targets = {
        "complexity": remaining * (45 / 81),
        "unit_tests": remaining * (36 / 81),
    }
    unit_factor = 1.0 if method_type == "全新开发" else 0.5
    scale = max(remaining, 1.0)
    candidates: list[tuple[float, float, float]] = []
    for complexity_hours in complexity_values:
        unit_hours = round(remaining - complexity_hours, 2)
        if unit_hours < -1e-9:
            continue
        unit_scenes = max(unit_hours, 0.0) / unit_factor
        score = (
            ((complexity_hours - targets["complexity"]) / scale) ** 2
            + ((unit_hours - targets["unit_tests"]) / scale) ** 2
            + abs(unit_scenes - round(unit_scenes)) * 0.02
        )
        candidates.append((score, complexity_hours, unit_scenes))
    if candidates:
        _score, complexity_hours, unit_scenes = min(candidates)
        values["codeComplexity"] = complexity_hours
        values["unitTestNumber"] = round(unit_scenes, 2)
    return values


def task_fingerprint(record: dict[str, Any]) -> tuple[str, ...]:
    return tuple(
        _text(record.get(field_name))
        for field_name in (
            "requireNo",
            "taskName",
            "developerId",
            "planStartDate",
            "planFinishDate",
            "module",
        )
    )


def _valid_iso_date(value: str) -> bool:
    try:
        date.fromisoformat(value)
        return True
    except (TypeError, ValueError):
        return False


def build_workload_preview(
    items: list[ExcelWorkItem],
    context: WorkloadContext,
    developer_filter: str = ALL_GROUP_MEMBERS,
) -> WorkloadPreview:
    rows: list[WorkloadPreviewRow] = []
    existing = {task_fingerprint(record) for record in context.existing_tasks}
    pending: set[tuple[str, ...]] = set()
    ignored_developers: set[str] = set()
    skipped = 0

    for item in items:
        if developer_filter != ALL_GROUP_MEMBERS and item.developer_name != developer_filter:
            skipped += 1
            continue
        if (
            developer_filter == ALL_GROUP_MEMBERS
            and item.developer_name
            and item.developer_name not in context.developers
        ):
            skipped += 1
            if item.developer_name:
                ignored_developers.add(item.developer_name)
            continue

        errors: list[str] = []
        warnings: list[str] = []
        requirement_description = context.requirements.get(item.require_no)
        developer = context.developers.get(item.developer_name)
        if not item.require_no:
            errors.append("需求编号为空")
        elif requirement_description is None:
            errors.append("当前版本中不存在该需求")
        if not item.task_name:
            errors.append("工作描述为空")
        if not developer:
            errors.append("责任人不属于当前分组")
        if not _valid_iso_date(item.plan_start_date):
            errors.append("计划开始时间格式无效")
        if not _valid_iso_date(item.plan_finish_date):
            errors.append("计划完成时间格式无效")
        if (
            _valid_iso_date(item.plan_start_date)
            and _valid_iso_date(item.plan_finish_date)
            and item.plan_start_date > item.plan_finish_date
        ):
            errors.append("计划开始时间晚于完成时间")
        if item.requested_days <= 0:
            errors.append("工时/天必须大于 0")

        if errors:
            rows.append(
                WorkloadPreviewRow(
                    source=item,
                    status="错误",
                    message="；".join(errors),
                    computed_hours=0.0,
                )
            )
            continue

        if context.ui_definition_path == SERVER_WORKLOAD_DEFINITION:
            allocation = allocate_server_workload(
                item.requested_days * 8,
                item.task_name,
                f"{item.require_no}|{item.task_name}",
            )
        else:
            allocation = allocate_ui_workload(
                item.requested_days * 8,
                item.task_name,
                f"{item.require_no}|{item.task_name}",
            )
        record: dict[str, Any] = {
            "rowId": uuid4().hex,
            "__isInserted": True,
            "isHandle": "0",
            "isChecked": False,
            "planVersion": context.plan_version,
            "module": context.module,
            "requireNo": item.require_no,
            "requireDescription": requirement_description,
            "taskName": item.task_name,
            "planStartDate": item.plan_start_date,
            "planFinishDate": item.plan_finish_date,
            "developerName": developer.name,
            "developerId": developer.user_id,
            "workLoad": None,
            "manDay": None,
            **allocation,
        }
        computed_hours = calculate_workload_hours(record)
        requested_hours = round(item.requested_days * 8, 2)
        if abs(computed_hours - requested_hours) > 0.01:
            warnings.append(
                f"按原规则计算为 {computed_hours / 8:g}天，"
                f"Excel 为 {item.requested_days:g}天"
            )
        fingerprint = task_fingerprint(record)
        if fingerprint in existing or fingerprint in pending:
            rows.append(
                WorkloadPreviewRow(
                    source=item,
                    status="已存在",
                    message="系统或本次 Excel 中已有相同任务，不会重复提交",
                    computed_hours=computed_hours,
                )
            )
            continue
        pending.add(fingerprint)
        rows.append(
            WorkloadPreviewRow(
                source=item,
                status="提醒" if warnings else "可提交",
                message="；".join(warnings) if warnings else "校验通过",
                computed_hours=computed_hours,
                record=record,
            )
        )

    return WorkloadPreview(
        rows=rows,
        source_row_count=len(items),
        skipped_row_count=skipped,
        ignored_developers=tuple(sorted(ignored_developers)),
    )


class WorkloadApiClient:
    """复用 BugApiClient 已登录会话访问工作量动态接口。"""

    def __init__(self, client: BugApiClient) -> None:
        self.client = client
        self._group_definitions: dict[str, str] = {}

    def _post_json(
        self,
        path: str,
        *,
        params: dict[str, Any] | None = None,
        data: dict[str, Any] | None = None,
        action: str,
    ) -> dict[str, Any]:
        if "Authorization" not in self.client.session.headers:
            raise AuthenticationError("请先登录")
        try:
            response = self.client.session.post(
                self.client._url(path),
                params=params,
                json=data,
                timeout=self.client.timeout,
            )
        except requests.RequestException as exc:
            raise BugClientError(f"{action}请求失败：{exc}") from exc
        if response.status_code == 401:
            raise AuthenticationError("登录状态已失效，请重新登录")
        try:
            payload = response.json()
        except ValueError as exc:
            raise BugClientError(f"{action}返回了无法识别的数据") from exc
        if response.status_code != 200 or payload.get("success") is False or payload.get(
            "result"
        ) == "ERROR":
            message = payload.get("message") or payload.get("msg") or f"{action}失败"
            raise BugClientError(str(message))
        return payload

    @staticmethod
    def _resolve_params(
        templates: list[dict[str, Any]] | None,
        plan_version: str,
        module: str,
    ) -> dict[str, Any]:
        values = {"planVersion": plan_version, "module": module}
        params: dict[str, Any] = {}
        for template in templates or []:
            if not isinstance(template, dict):
                continue
            for key, source in template.items():
                params[key] = values.get(str(source), source)
        return params

    def get_groups(self) -> list[WorkloadGroup]:
        payload = self._post_json(
            "/rest/uiDef/getWorkItemGroup",
            action="读取工作量分组",
        )
        groups: list[WorkloadGroup] = []
        stack = [payload.get("data")]
        while stack:
            node = stack.pop()
            if not isinstance(node, dict):
                continue
            for method in node.get("methods") or []:
                definition_path = method.get("getUiDef") if isinstance(method, dict) else None
                if definition_path in SUPPORTED_WORKLOAD_DEFINITIONS:
                    name = _text(node.get("workItemGroup"))
                    if name:
                        groups.append(WorkloadGroup(name, definition_path))
            stack.extend(reversed(node.get("children") or []))
        self._group_definitions = {
            group.name: group.ui_definition_path for group in groups
        }
        return groups

    def _definition_path(self, module: str) -> str:
        cached = self._group_definitions.get(module)
        if cached:
            return cached
        return (
            UI_WORKLOAD_DEFINITION
            if module.endswith("-前端")
            else SERVER_WORKLOAD_DEFINITION
        )

    def _get_definition(self, module: str) -> tuple[str, dict[str, Any]]:
        definition_path = self._definition_path(module)
        definition = self._post_json(
            f"/rest{definition_path}",
            action="读取工作量表定义",
        ).get("data") or {}
        if not isinstance(definition, dict):
            raise BugClientError("工作量表定义格式异常")
        return definition_path, definition

    def _developer_binding(
        self,
        definition: dict[str, Any],
        plan_version: str,
        module: str,
    ) -> tuple[str, dict[str, Any]]:
        for column in definition.get("columns") or []:
            if not isinstance(column, dict) or column.get("fieldName") != "developerName":
                continue
            binding = column.get("bind") or {}
            return (
                _text(binding.get("Interface")),
                self._resolve_params(binding.get("params"), plan_version, module),
            )
        return "", {}

    @staticmethod
    def _parse_developers(data: Any, module: str) -> dict[str, Developer]:
        developers: dict[str, Developer] = {}
        for item in data if isinstance(data, list) else []:
            if not isinstance(item, dict):
                continue
            name = _text(item.get("name"))
            user_id = _text(item.get("userId"))
            job = _text(item.get("job"))
            if not name or not user_id:
                continue
            if module.endswith("-前端") and job and job != "前端":
                continue
            if not module.endswith("-前端") and job and job != "后端":
                continue
            developers[name] = Developer(name, user_id, job)
        return developers

    def get_developers(self, module: str) -> list[Developer]:
        _definition_path, definition = self._get_definition(module)
        path, params = self._developer_binding(definition, "", module)
        if not path:
            raise BugClientError("工作量表定义中缺少开发资源接口")
        data = self._post_json(
            f"/rest{path}",
            params=params,
            action="读取开发资源",
        ).get("data") or []
        return sorted(self._parse_developers(data, module).values(), key=lambda item: item.name)

    def get_context(self, plan_version: str, module: str) -> WorkloadContext:
        definition_path, definition = self._get_definition(module)
        load_path = ""
        save_path = ""
        load_params: dict[str, Any] = {}
        for method in definition.get("methods") or []:
            if not isinstance(method, dict):
                continue
            if method.get("loadData"):
                load_path = _text(method.get("loadData"))
                load_params = self._resolve_params(
                    method.get("params"), plan_version, module
                )
            if method.get("saveData"):
                save_path = _text(method.get("saveData"))
        if not load_path or not save_path:
            raise BugClientError("工作量表定义中缺少加载或保存接口")

        requirement_path = ""
        requirement_params: dict[str, Any] = {}
        for binding in definition.get("bind") or []:
            if isinstance(binding, dict) and binding.get("id") == "RequiresByPlan":
                requirement_path = _text(binding.get("Interface"))
                requirement_params = self._resolve_params(
                    binding.get("params"), plan_version, module
                )

        developer_path, developer_params = self._developer_binding(
            definition, plan_version, module
        )
        if not requirement_path or not developer_path:
            raise BugClientError("工作量表定义中缺少需求或开发资源接口")

        requirements_data = self._post_json(
            f"/rest{requirement_path}",
            params=requirement_params,
            action="读取版本需求",
        ).get("data") or []
        developer_data = self._post_json(
            f"/rest{developer_path}",
            params=developer_params,
            action="读取开发资源",
        ).get("data") or []
        existing_tasks = self._post_json(
            f"/rest{load_path}",
            params=load_params,
            action="读取已有工作量",
        ).get("data") or []

        requirements = {
            _text(item.get("requireNo")): _text(item.get("requireDescription"))
            for item in requirements_data
            if isinstance(item, dict) and _text(item.get("requireNo"))
        }
        developers = self._parse_developers(developer_data, module)
        existing_list = [item for item in existing_tasks if isinstance(item, dict)]
        ord_values = [
            int(item["ord"])
            for item in existing_list
            if isinstance(item.get("ord"), (int, float))
        ]
        return WorkloadContext(
            plan_version=plan_version,
            module=module,
            load_path=load_path,
            save_path=save_path,
            load_params=load_params,
            requirements=requirements,
            developers=developers,
            existing_tasks=existing_list,
            next_ord=max(ord_values, default=-1) + 1,
            ui_definition_path=definition_path,
        )

    def submit(
        self,
        context: WorkloadContext,
        rows: list[WorkloadPreviewRow],
    ) -> WorkloadSubmitResult:
        records: list[dict[str, Any]] = []
        for index, row in enumerate(rows):
            if not row.is_submittable or row.record is None:
                continue
            record = dict(row.record)
            record["ord"] = context.next_ord + index
            records.append(record)
        if not records:
            raise BugClientError("没有可提交的工作量记录")

        payload = self._post_json(
            f"/rest{context.save_path}",
            data={
                "insertRecords": records,
                "updateRecords": [],
                "removeRecords": [],
            },
            action="批量保存工作量",
        )
        refreshed = self._post_json(
            f"/rest{context.load_path}",
            params=context.load_params,
            action="回查工作量",
        ).get("data") or []
        refreshed_records = [item for item in refreshed if isinstance(item, dict)]
        refreshed_fingerprints = {
            task_fingerprint(item) for item in refreshed_records
        }
        verified = sum(
            task_fingerprint(record) in refreshed_fingerprints for record in records
        )
        message = _text(payload.get("message")) or "保存成功"
        return WorkloadSubmitResult(
            submitted_count=len(records),
            verified_count=verified,
            before_count=len(context.existing_tasks),
            after_count=len(refreshed_records),
            message=message,
        )
